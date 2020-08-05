import os
import cv2
import matplotlib.pyplot as plt
import numpy as np
import utils
import shutil
import re
import configparser
from openpyxl import load_workbook

def rename(directory):
    '''Rename directory
    @directory [in]:directory to be renamed
    '''
    index = 1
    #left_dir = directory + '\\left\\'
    filelist = os.listdir(directory)
    for filename in filelist:
        print(filename)
        os.rename(os.path.join(directory, filename), os.path.join(directory, str(index) + '.JPG'))
        index += 1
    # index = 1
    # right_dir = directory + '\\right\\'
    # filelist = os.listdir(right_dir)
    # for filename in filelist:
    #     print(filename)
    #     os.rename(os.path.join(right_dir, filename), os.path.join(right_dir, str(index) + '.JPG'))
    #     index += 1

def resize(old_dir, new_dir, dim=(1920, 1280)):
    '''Resize images in specified directory
    @old_dir [in]: old directory which contains images to be resized
    @new_dir [in]: directory to save resized images
    @dim [in]: new size of images
    '''
    filelist = os.listdir(old_dir)
    for filename in filelist:
        img = cv2.imread(os.path.join(old_dir, filename), cv2.IMREAD_UNCHANGED)
        resize = cv2.resize(img, dim)
        cv2.imwrite(os.path.join(new_dir, filename), resize)

# def show(left_dir, right_dir):
#     '''Show images in directories by pair left - right
#     @left_dir [in]: directory which contains left image
#     @left_dir [in]: directory which contains right image
#     '''
#     left = cv2.imread('Img_2\\left\\1.JPG')
#     right = cv2.imread('Img_2\\right\\1.JPG')
#     left = cv2.cvtColor(left, cv2.COLOR_RGB2BGR)
#     right = cv2.cvtColor(right, cv2.COLOR_RGB2BGR)
#     plt.figure() 
#     plt.axis("off")
#     ax1 = plt.subplot(1, 2, 1)
#     ax2 = plt.subplot(1, 2, 2, sharex= ax1, sharey=ax1)
#     ax1.imshow(left)#, cmap=cv2.COLOR_RGB2BGR)
#     ax2.imshow(right)#, cmap=cv2.COLOR_RGB2BGR)
#     plt.show()

def extractVidToFrame(vidUrl, limitLen=False):
    '''
    Extract video into frames
    @vidUrl [in]: the path to the video to be extracted
    @limitLen [in]: when True, just get the frames from first 30 seconds
    '''
    vid = cv2.VideoCapture(vidUrl)
    idx = vidUrl.find('.')
    fps = vid.get(cv2.CAP_PROP_FPS)
    prefix = vidUrl[ : idx]
    if not os.path.isdir('Video\\Output_' + prefix):
        os.makedirs('Video\\Output_' + prefix)
    success, img = vid.read()
    count = 1
    while success:
        if success == True and ((count < (fps * 31) and limitLen is True) or limitLen is False):
            cv2.imwrite('Video\\Output_' + prefix + '\\' + str(count) + '.jpg', img)
            print('cnt: ', count)
            count = count + 1
        success, img = vid.read()
    print('Extract done!') 

def reduceCell(fileName, sId=3, col=15):
    '''
    Delete cells which missed value(s)
    @fileName [in]: the path to the excel file
    @sId [in]: sheet index
    @col [in]: column M
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sId]
    #limit = 3960
    i = 2
    print('Reduce begin')
    while sheet.cell(row=i, column=1).value is not None:
    # print(i, float(sheet.cell(row=i, column=5).value))
    # print(i, float(sheet.cell(row=i, column=6).value))
    # print(i, float(sheet.cell(row=i, column=7).value))
    # print(i, float(sheet.cell(row=i, column=10).value))
    # print(i, float(sheet.cell(row=i, column=11).value))
    # print(i, float(sheet.cell(row=i, column=12).value))
    # sheet.cell(row=i, column=15).value = float(sheet.cell(row=i, column=5).value) * float(sheet.cell(row=i, column=6).value) / float(sheet.cell(row=i, column=7).value)
    # sheet.cell(row=i, column=16).value = float(sheet.cell(row=i, column=10).value) * float(sheet.cell(row=i, column=11).value) / float(sheet.cell(row=i, column=12).value)
    # i = i + 1
        if sheet.cell(row=i, column=col).value is None:# and sheet.cell(row=i, column=1).value is not None:
            print('Erase (' + str(sheet.cell(row=i, column=1).value) + ', ' + str(sheet.cell(row=i, column=2).value)+ ', ' + str(sheet.cell(row=i, column=3).value) + ')') 
            sheet.delete_rows(i)
        else:#if sheet.cell(row=i, column=15).value is not None and sheet.cell(row=i, column=1) is not None:
            print('Keep (' + str(sheet.cell(row=i, column=1).value) + ', ' + str(sheet.cell(row=i, column=2).value)+ ', ' + str(sheet.cell(row=i, column=3).value) + ')') 
            i = i + 1
        # if sheet.cell(row=i, column=1).value is None:
        #     print('End game!')
        #     break
    print('Reduce finished. Saving...')
    wb.save(fileName)
    print('Reduce finished. Saved!')

def copyCell(fileName, originId=2, copyId=3):
    '''
    Copy content from one sheet to another sheet (remove empty rows)
    @fileName [in]: the path to the excel file
    @originId [in]: index of the original sheet
    @copyId [in]: index of the copy sheet
    '''
    wb = load_workbook(fileName)
    origin_sheet = wb.worksheets[originId]
    copy_sheet = wb.worksheets[copyId]
    i = 1
    while origin_sheet.cell(row=i, column=1).value is not None:
        if origin_sheet.cell(row=i, column=15).value is not None:
            print(origin_sheet[i])
            copy_sheet.append(origin_sheet[i])
        i = i + 1
    wb.save(fileName)

def createRangeColumn(fileName, sId=3):
    '''
    Create "Size ratio in range" column value
    @fileName [in]: excel file
    @sId [in]: sheet index
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sId]
    i = 2
    while sheet.cell(row=i, column=1).value is not None:
        if float(sheet.cell(row=i, column=24).value) <= 2:
            sheet.cell(row=i, column=25).value = '1'
        else:
            sheet.cell(row=i, column=25).value = '0'
        i = i + 1 
    wb.save(fileName)   

def createRatioColumn(fileName, sId=3):
    '''
    Create "Width ratio", "Height ratio", "Width ratio in range", "Height ratio in range" column value
    @fileName [in]: excel file
    @sId [in]: sheet index
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sId]
    i = 2
    while sheet.cell(row=i, column=1).value is not None:
        sheet.cell(row=i, column=26).value = float(sheet.cell(row=i, column=5).value) / float(sheet.cell(row=i, column=10).value)
        sheet.cell(row=i, column=27).value = float(sheet.cell(row=i, column=6).value) / float(sheet.cell(row=i, column=11).value)
        if float(sheet.cell(row=i, column=26).value) < 2:
            sheet.cell(row=i, column=28).value = '1'
        else:
            sheet.cell(row=i, column=28).value = '0'
        if float(sheet.cell(row=i, column=27).value) < 2:
            sheet.cell(row=i, column=29).value = '1'
        else:
            sheet.cell(row=i, column=29).value = '0'
        i = i + 1 
    wb.save(fileName) 

def removeBus(fileName, stereoIdx=-1, pairingIdx_left=-1, pairingIdx_right=-1):
    '''
    Remove bus records in dataset
    @fileName [in]: excel file
    @stereoIdx [in]: sheet index of stereo matching dataset
    @pairingIdx_left [in]: sheet index of taillight pairing in left image dataset
    @pairingIdx_right [in]: sheet index of taillight pairing in right image dataset
    '''
    wb = load_workbook(fileName)

    '''Stereo matching dataset'''
    if stereoIdx != -1:
        stereoSheet = wb.worksheets[stereoIdx]
        i = 2
        while stereoSheet.cell(row=i, column=1).value is not None:
            img = str(stereoSheet.cell(row=i, column=1).value)
            stereo_left = str(stereoSheet.cell(row=i, column=2).value)
            stereo_right = str(stereoSheet.cell(row=i, column=3).value)
            if img in utils.bus_map_left and img in utils.bus_map_right and ((stereo_left in utils.bus_map_left[img]) or (stereo_right in utils.bus_map_right[img])):
                print('Erase (' + img + ', ' + stereo_left + ', ' + stereo_right + ')') 
                stereoSheet.delete_rows(i)
            else:
                print('Keep (' + img + ', ' + stereo_left + ', ' + stereo_right + ')') 
                i = i + 1

    '''Taillight pairing in left image dataset'''
    if pairingIdx_left != -1:
        pairingleftSheet = wb.worksheets[pairingIdx_left]
        i = 2
        while pairingleftSheet.cell(row=i, column=1).value is not None:
            img = str(pairingleftSheet.cell(row=i, column=1).value)
            pairing_left = str(pairingleftSheet.cell(row=i, column=2).value)
            pairing_right = str(pairingleftSheet.cell(row=i, column=3).value)
            if img in utils.bus_map_left and ((pairing_left in utils.bus_map_left[img]) or (pairing_right in utils.bus_map_left[img])):
                print('Erase (' + img + ', ' + pairing_left+ ', ' + pairing_right + ')') 
                pairingleftSheet.delete_rows(i)
            else:
                print('Keep (' + img + ', ' + pairing_left + ', ' + pairing_right + ')') 
                i = i + 1   

    '''Taillight pairing in right image dataset'''
    if pairingIdx_right != -1:
        pairingrightSheet = wb.worksheets[pairingIdx_right]
        i = 2
        while pairingrightSheet.cell(row=i, column=1).value is not None:
            img = str(pairingrightSheet.cell(row=i, column=1).value)
            pairing_left = str(pairingrightSheet.cell(row=i, column=2).value)
            pairing_right = str(pairingrightSheet.cell(row=i, column=3).value)
            if img in utils.bus_map_right and ((pairing_left in utils.bus_map_right[img]) or (pairing_right in utils.bus_map_right[img])):
                print('Erase (' + img + ', ' + pairing_left + ', ' + pairing_right + ')') 
                pairingrightSheet.delete_rows(i)
            else:
                print('Keep (' + img + ', ' + pairing_left + ', ' + pairing_right + ')') 
                i = i + 1  

    wb.save(fileName)     
    
    # origin_sheet = wb.worksheets[0]
    # copy_sheet = wb.worksheets[3]
    # i = 1
    # while origin_sheet.cell(row=i, column=1).value is not None:
    #     if origin_sheet.cell(row=i, column=15).value is not None:
    #         copy_sheet.append(origin_sheet[i])
    #     i = i + 1

def parseGroundtruthData(iniFile):
    '''
    Parse ground truth data from INI file to use in code:
    @iniFile [in]: INI file name
    '''
    stereo_matching_left = dict()
    stereo_matching_right = dict()
    taillight_pairing_left = dict()
    taillight_pairing_right = dict()
    config = configparser.ConfigParser(allow_no_value=True)
    config.read(iniFile)
    imgs = config.sections()
    for img in imgs:
        #print(str(img))
    # for i in range(1, numOfImg + 1):
    #     img = str(i) + '.JPG'
        stereo_matching_left[img] = re.split('[,]', config[img]['stereo_matching_left'])
        stereo_matching_right[img] = re.split('[,]', config[img]['stereo_matching_right'])
        taillight_pairing_left[img] = re.split('[,]', config[img]['taillight_pairing_left'])
        taillight_pairing_right[img] = re.split('[,]', config[img]['taillight_pairing_right'])
        # if len(taillight_pairing_left[img]) != 0: 
        #     print(taillight_pairing_left[img])
        #print(np.shape(stereo_matching_left[img]))
    
    return imgs, stereo_matching_left, stereo_matching_right, taillight_pairing_left, taillight_pairing_right

def generateStereoDataset(fileName, stereoIdx, col, stereo_matching_left, stereo_matching_right, imgs = [], reduceMatch=False):
    '''
    Generate stereo matching dataset:
    @fileName [in]: excel file,
    @stereoIdx [in]: sheet index of stereo matching dataset,
    @col [in]: column 'Match' in dataset,
    @stereo_matching_left [in]: list of stereo matching taillights in left img,
    @stereo_matching_right [in]: list of stereo matching taillights in right img,
    @imgs [in]: list of image in ground truth data (default is empty []),
    @reduceMatch [in]: if yes: reduce the number of possible stereo matching case (in case 1 in 2 taillights is already matched) (default is False)
    '''
    wb = load_workbook(fileName)
    stereoSheet = wb.worksheets[stereoIdx]
    i = 2
    print('Begin matching') 
    while stereoSheet.cell(row=i, column=1).value is not None:
        img = str(stereoSheet.cell(row=i, column=1).value)
        stereo_left = str(stereoSheet.cell(row=i, column=2).value)
        stereo_right = str(stereoSheet.cell(row=i, column=3).value)
        #img not exist in groundtruth dataset'''
        # if (imgs != []) or (img not in imgs):
        #     print('Erase ' + img + ', ' + stereo_left + ', ' + stereo_right) 
            #stereoSheet.delete_rows(i) 
        if(img in stereo_matching_left) and (img in stereo_matching_right):  #img exists in groundtruth dataset'''
            if (stereo_left in stereo_matching_left[img]) and (stereo_right in stereo_matching_right[img]):#stereo matching exists'''
                idl = stereo_matching_left[img].index(stereo_left)
                idr = stereo_matching_right[img].index(stereo_right)         
                if idl == idr:
                    print('Exist 1')
                    stereoSheet.cell(row=i, column=col).value = '1'
                else:
                    if reduceMatch is True:
                        print('Exist but must erase (' + img + ', ' + stereo_left + ', ' + stereo_right + ')') 
                        # stereoSheet.delete_rows(i)
                        # i = i - 1
                    else:
                        print('Exist 0. 1')
                        stereoSheet.cell(row=i, column=col).value = '0'
                #i = i + 1
            elif (stereo_left in stereo_matching_left[img]) or (stereo_right in stereo_matching_right[img]):
                if reduceMatch is True:
                    print('Exist 0. 1')
                    stereoSheet.cell(row=i, column=col).value = '0'
            else:   #stereo matching not exist'''
                print('Not exist so must erase (' + img + ', ' + stereo_left + ', ' + stereo_right + ')') 
                #stereoSheet.delete_rows(i)
        else:
            print('Erase ' + img + ', ' + stereo_left + ', ' + stereo_right)
        #wb.save(fileName)
        i = i + 1
    print('Finished matching. Saving...')
    wb.save(fileName)  
    print('Finished matching. Saved!')   

def generatePairingDataset(fileName, pairingIdx, col, stereo_matching, taillight_pairing, imgs = [], reducePair=False):
    '''
    Generate taillight pairing dataset:
    @fileName [in]: excel file,
    @pairingIdx [in]: sheet index of taillight pairing dataset,
    @col [in]: column 'Match' in dataset,
    @stereo_matching [in]: list of stereo matching taillights,
    @taillight_pairing [in]: list of paired taillights,
    @stereo [in]: decide which image side to be used for generating (left image/right image),
    @imgs [in]: list of image in ground truth data (default is empty [])
    @reducePair [in]: if yes: reduce the number of potential pair (in case 1 in 2 taillights is already paired) (default is False)
    '''
    wb = load_workbook(fileName)
    pairingSheet = wb.worksheets[pairingIdx]
    i = 2
    print('Begin pairing') 
    while pairingSheet.cell(row=i, column=1).value is not None:
        img = str(pairingSheet.cell(row=i, column=1).value)
        print(img)
        pairing_left = str(pairingSheet.cell(row=i, column=2).value)
        pairing_right = str(pairingSheet.cell(row=i, column=3).value)
        #img not exist in groundtruth dataset'''
        # if (imgs != []) or (img not in imgs):
        #     print('Erase ' + img + ', ' + pairing_left + ', ' + pairing_right) 
            #pairingSheet.delete_rows(i) 
        if(img in taillight_pairing) or (img in stereo_matching):   #img exists in groundtruth dataset'''
            if (img in taillight_pairing) and (pairing_left in taillight_pairing[img]) and (pairing_right in taillight_pairing[img]):#taillight pair exists'''
                idl = taillight_pairing[img].index(pairing_left)
                idr = taillight_pairing[img].index(pairing_right)
                if (idl + 1 == idr) and (idl & 1 == 0):
                    print('Exist 1')
                    pairingSheet.cell(row=i, column=col).value = '1'
                else:
                    if reducePair is True: #and idl + 1 < idr:
                        print('Exist but must erase (' + img + ', ' + pairing_left + ', ' + pairing_right + ')') 
                        # pairingSheet.delete_rows(i)
                        # i = i - 1
                    else:
                        print('Exist 0. 1')
                        pairingSheet.cell(row=i, column=col).value = '0'
                #i = i + 1           
            # elif (pairing_left in stereo_matching[img]) and (pairing_right in stereo_matching[img]): #taillight pair not exist but stereo matching exist'''
            #     pairingSheet.cell(row=i, column=col).value = '0'
                #i = i + 1
            elif ((img in stereo_matching and pairing_left in stereo_matching[img])\
                or (img in taillight_pairing and pairing_left in taillight_pairing[img]))\
                and ((img in stereo_matching and pairing_right in stereo_matching[img])\
            or (img in taillight_pairing and pairing_right in taillight_pairing[img])):                #if reducePair is True:
                    #pairingSheet.cell(row=i, column=col).value = '0'
                print('Exist 0. 2')
                pairingSheet.cell(row=i, column=col).value = '0'
            else:
                print('Not exist so must erase (' + img + ', ' + pairing_left + ', ' + pairing_right + ')') 
                #pairingSheet.delete_rows(i)  
        #wb.save(fileName)
        else:
            print('Erase ' + img + ', ' + pairing_left + ', ' + pairing_right) 
        i = i + 1 
    print('Finished pairing. Saving...')         
    wb.save(fileName)
    print('Finished pairing. Saved!') 

def normalizeDistance(fileName, sheetId, colId, dbl=-1, dbr=-1, getAbs=False, imgWidth=0, limit=-1):
    '''
    Normalize inter-distance:
    @fileName [in]: excel file,
    @sheetId [in]: sheet index,
    @col [in]: column 'inter_distance' index,
    @dbl [in]: column 'DBL' index (default is -1),
    @dbr [in]: column 'DBR' index (default is -1),
    @getAbs [in]: get absolute value or not (default is False),
    @imgWidth [in]: image width (default is 0),
    @limit [in]: set a size limitation for the dataset
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    print('Begin normalizing')
    while sheet.cell(row=i, column=1).value is not None:
        if limit != -1:
            number,_ = os.path.splitext(sheet.cell(row=i, column=1).value)
            if limit < int(number):
                i = i + 1
                continue
        if dbl != -1 and dbr != -1:
            left = float(sheet.cell(row=i, column=dbl).value)
            right = float(sheet.cell(row=i, column=dbr).value)
            sheet.cell(row=i, column=colId).value = str(left - right)
        if imgWidth != 0:
            sheet.cell(row=i, column=colId).value = str(float(sheet.cell(row=i, column=colId).value) / imgWidth)            
        if getAbs is True:
            # oldValue = float(sheet.cell(row=i, column=colId).value)
            # sheet.cell(row=i, column=colId).value = str(oldValue / imgWidth)
            sheet.cell(row=i, column=colId).value = str(abs(float(sheet.cell(row=i, column=colId).value)))
        #     else:
        #         sheet.cell(row=i, column=colId).value = str(left - right)
        # if imgWidth != 0:
        #     sheet.cell(row=i, column=colId).value = str(float(sheet.cell(row=i, column=colId).value) / imgWidth)
        print('Row ' + str(i) + ': ' + str(sheet.cell(row=i, column=colId).value))
        i = i + 1
    print('Finished normalizing. Saving...')
    wb.save(fileName)
    print('Finished normalizing. Saved!')

def verticalDistance(fileName, sheetId, colRes, colId_left=-1, colId_right=-1, getAbs=False, imgHeight=0):
    '''
    Normalize vertical-distance:
    @fileName [in]: excel file,
    @sheetId [in]: sheet index,
    @colRes [in]: column index Vertical distance,
    @colId_left [in]: column index VL (default is -1),
    @colId_right [in]: column index VR (default is -1),
    @getAbs [in]: get absolute value or not (default is False),
    @imgHeight [in]: image height (default is 0)
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    print('Begin verticalizing')
    while sheet.cell(row=i, column=1).value is not None:
        if colId_left !=-1 and colId_right !=-1:
            vl = float(sheet.cell(row=i, column=colId_left).value)
            vr = float(sheet.cell(row=i, column=colId_right).value)
            sheet.cell(row=i, column=colRes).value = str(vl - vr)
        if getAbs is True:
            sheet.cell(row=i, column=colRes).value = str(abs(float(sheet.cell(row=i, column=colRes).value)))
            #     print('Row ' + str(i) + ': ' + str(sheet.cell(row=i, column=colRes).value))
            # else:
            #     sheet.cell(row=i, column=colRes).value = str(vl - vr)
        if imgHeight != 0:
            sheet.cell(row=i, column=colRes).value = str(float(sheet.cell(row=i, column=colRes).value) / imgHeight)
        i = i + 1
    print('Finish verticalizing. Saving...')
    wb.save(fileName)
    print('Finish verticalizing. Saved!')

def getDifference(fileName, sheetId, col_DBL, col_DBR, col_DBD, col_VL, col_VR, col_VD):
    '''
    Calculate Distance to Border & Vertical Position differences
    @fileName [in]: excel file
    @sheetId [in]: sheet index
    @col_DBL [in]: column index DBL
    @col_DBR [in]: column index DBR
    @col_DBD [in]: column index DBD (DB difference)
    @col_VL [in]: column index VL
    @col_VR [in]: column index VR
    @col_VD [in]: column index VD (V difference)
    '''    
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    while sheet.cell(row=i, column=1).value is not None:
        dbl = float(sheet.cell(row=i, column=col_DBL).value)
        dbr = float(sheet.cell(row=i, column=col_DBR).value)
        sheet.cell(row=i, column=col_DBD).value = str(dbl - dbr)
        vl = float(sheet.cell(row=i, column=col_VL).value)
        vr = float(sheet.cell(row=i, column=col_VR).value)
        sheet.cell(row=i, column=col_VD).value = str(vl - vr)
        print('Row ' + str(i) + ': ' + sheet.cell(row=i, column=col_DBD).value + ', ' + sheet.cell(row=i, column=col_VD).value)
        i = i + 1
    wb.save(fileName)

def rename(origin_dir, new_dir, offset):
    '''
    Rename image files from folder
    @origin_dir [in]: origin image folder
    @new_dir [in]: new image folder
    @offset [in]: offset to rename image
    '''
    filelist = os.listdir(origin_dir)
    for filename in filelist:
        new_name = int(os.path.splitext(filename)[0]) - offset
        if new_name > 0:
            print('Old name: ', str(filename))
            original = os.path.join(origin_dir, filename)
            target = os.path.join(new_dir, str(new_name) + '.jpg')
            shutil.copyfile(original, target)
            print('New name: ' + str(new_name) + '.jpg')
            print('---')

def normalizeSizeRatio(fileName, sheetId, sizeLeftId, sizeRightId, ratioId, absRatioId, limit=-1):
    '''
    Normalize size ratio:
    @fileName [in]: excel file name,
    @sheetId [in]: sheet Id,
    @sizeLeftId [in]: column index of left size,
    @sizeRightId [in]: column index of right size,
    @ratioId[in]: column index of size ratio,
    @absRatioId[in]: column index of absolute size ratio,
    @limit [in]: set a size limitation for the dataset
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    print('Begin calculating ratio')
    while sheet.cell(row=i, column=1).value is not None:
        if limit != -1:
            number,_ = os.path.splitext(sheet.cell(row=i, column=1).value)
            if limit < int(number):
                i = i + 1
                continue
        sl = float(sheet.cell(row=i, column=sizeLeftId).value)
        sr = float(sheet.cell(row=i, column=sizeRightId).value)
        size_ratio_abs = min(sl, sr) / max(sl, sr)
        if sl < sr:
            size_ratio = -size_ratio_abs
        else:
            size_ratio = size_ratio_abs
        sheet.cell(row=i, column=absRatioId).value = str(size_ratio_abs)
        sheet.cell(row=i, column=ratioId).value = str(size_ratio)
        i = i + 1
    print('Finish calculating ratio. Saving...')
    wb.save(fileName)
    print('Finish calculating ratio. Saved!')
# wb = load_workbook('taillight_pairing_new.xlsx')
# sheet = wb.worksheets[0]
# i = 2
# print('Begin verticalizing')
# while sheet.cell(row=i, column=1).value is not None:
#     sheet.cell(row=i, column=14).value = sheet.cell(row=i, column=12).value
#     sheet.cell(row=i, column=12).value = str(abs(float(sheet.cell(row=i, column=12).value)))
#     i = i + 1
# print('Finish verticalizing. Saving...')
# wb.save('taillight_pairing_new.xlsx')
# print('Finish verticalizing. Saved!')

def normalizeAreaRatio(fileName, sheetId, widthLeftId, widthRightId, heightLeftId, heightRightId, ratioId, absRatioId):
    '''
    Normalize size ratio:
    @fileName [in]: excel file name,
    @sheetId [in]: sheet Id,
    @widthLeftId [in]: column index of left width,
    @widthRightId [in]: column index of right width,
    @heightLeftId [in]: column index of left height,
    @heightRightId [in]: column index of right height,
    @ratioId[in]: column index of size ratio,
    @absRatioId[in]: column index of absolute size ratio
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    print('Begin calculating area')
    while sheet.cell(row=i, column=1).value is not None:
        sl = float(sheet.cell(row=i, column=widthLeftId).value) * float(sheet.cell(row=i, column=heightLeftId).value)
        sr = float(sheet.cell(row=i, column=widthRightId).value) * float(sheet.cell(row=i, column=heightRightId).value)
        size_ratio_abs = min(sl, sr) / max(sl, sr)
        if sl < sr:
            size_ratio = -size_ratio_abs
        else:
            size_ratio = size_ratio_abs
        sheet.cell(row=i, column=absRatioId).value = str(size_ratio_abs)
        sheet.cell(row=i, column=ratioId).value = str(size_ratio)
        i = i + 1
    print('Finish calculating area. Saving...')
    wb.save(fileName)
    print('Finish calculating area. Saved!')
def updateDistance(fileName, sheetId, distanceId, sizeLeftId, sizeRightId):
    '''
    Update distane:
    @fileName [in]: excel file name,
    @sheetId [in]: sheet Id,
    @distanceId [in]: column index of distance,
    @sizeLeftId [in]: column index of left size,
    @sizeRightId [in]: column index of right size
    '''
    wb = load_workbook(fileName)
    sheet = wb.worksheets[sheetId]
    i = 2
    print('Begin updating distance')
    while sheet.cell(row=i, column=1).value is not None:
        sl = float(sheet.cell(row=i, column=sizeLeftId).value)
        sr = float(sheet.cell(row=i, column=sizeRightId).value)
        d = float(sheet.cell(row=i, column=distanceId).value) + (sr - sl) / 2
        sheet.cell(row=i, column=distanceId).value = str(d)
        i = i + 1
    print('Finish updating distance. Saving...')
    wb.save(fileName)
    print('Finish updating distance. Saved!')    
#reduceCell('Set_2\\stereo_matching_without_bus.xlsx', 0, 13)

'''Initialize ground truth dataset'''
#imgs, stereo_matching_left, stereo_matching_right, taillight_pairing_left, taillight_pairing_right = parseGroundtruthData('Set_2\\Set_2.ini')
# print(np.shape(imgs))
#print('Initialize done!')
# print(imgs)
# print(np.shape(stereo_matching_left))
# stereo_matching_left = utils.stereo_matching_left
# stereo_matching_right = utils.stereo_matching_right
# # taillight_pairing_left = utils.taillight_pairing_left
# taillight_pairing_right = utils.taillight_pairing_right

'''Generate stereo matching dataset'''
# generateStereoDataset('Img_2\\stereo_matching_without_bus.xlsx', 0, 13, stereo_matching_left, stereo_matching_right)
#generateStereoDataset('Set_2\\stereo_matching_without_bus.xlsx', 0, 13, stereo_matching_left, stereo_matching_right, imgs)
normalizeDistance('Set_2\\stereo_matching_without_bus.xlsx', 0, 14, getAbs=True)
verticalDistance('Set_2\\stereo_matching_without_bus.xlsx', 0, 15, getAbs=True)
# normalizeDistance('Img_2\\stereo_matching_without_bus.xlsx', 0, 10, 14, 16)
# verticalDistance('Img_2\\stereo_matching_without_bus.xlsx', 0, 11, 15, 17)

'''Generate taillight pairing dataset on left imgs'''
#generatePairingDataset('Set_2\\pairing_left.xlsx', 0, 13, stereo_matching_left, taillight_pairing_left, imgs)#, True)
# verticalDistance('Set_2\\taillight_pairing.xlsx', 0, 6, 9, 12)
# normalizeDistance('Set_2\\taillight_pairing.xlsx', 0, 11, imgWidth=1920)

#stereo_matching_left = utils.stereo_matching_left
#stereo_matching_right = utils.stereo_matching_right
#taillight_pairing_left = utils.taillight_pairing_left
# generateStereoDataset('Img_2\\stereo_matching_without_bus.xlsx', 0, 17, stereo_matching_left, stereo_matching_right)#, True)
# removeBus('Img_2\\stereo_matching_without_bus.xlsx', 0)
# generatePairingDataset('Img_2\\taillight_pairing.xlsx', 0, 13, stereo_matching_left, taillight_pairing_left)#, True)
# removeBus('Img_2\\taillight_pairing.xlsx', 0)
# normalizeDistance('Img_2\\taillight_pairing.xlsx', 0, 11, imgWidth=2464)
'''Generate taillight pairing dataset on right imgs'''
#generatePairingDataset('Set_2\\pairing_right.xlsx', 0, 13, stereo_matching_right, taillight_pairing_right, imgs)#, True)

#generatePairingDataset('Img_2\\pairing_right.xlsx', 0, 12, stereo_matching_right, taillight_pairing_right)
# verticalDistance('Img_2\\pairing_right.xlsx', 0, 6, 9, 13, 1)

# reduceCell('Img_2\\left\\pairing_left.xlsx', 0, 13)
# reduceCell('Img_2\\right\\pairing_right.xlsx', 0, 12)
#reduceCell('Img_2\\stereo_matching_without_bus.xlsx', 0, 13)
# generatePairingDataset('Set_2\\pairing_right.xlsx', 0, 13, stereo_matching_right, taillight_pairing_right, imgs)
# normalizeDistance('Set_2\\pairing_right.xlsx', 0, 11, imgWidth=1920)
# verticalDistance('Set_2\\pairing_right.xlsx', 0, 6, 9, 12)
# reduceCell('Set_2\\pairing_left_reduce.xlsx', 0, 13)
# reduceCell('Set_2\\pairing_right_reduce.xlsx', 0, 13)
# reduceCell('Set_2\\stereo_matching_without_bus_reduce.xlsx', 0, 13)
#normalizeDistance('Img_2\\stereo_matching_without_bus.xlsx', 0, 10, 14, 15, 1)
#verticalDistance('Img_2\\stereo_matching_without_bus.xlsx', 0, 16, 17, 11, 1)






'''
# extractVidToFrame('MVI_1308.MOV')
# extractVidToFrame('MVI_1310.MOV')
# extractVidToFrame('DSC_8860.MOV')
# extractVidToFrame('DSC_8862.MOV')
#rename('Video\\Set 2\\right_bak', 'Video\\Set 2\\right', 26)
# generateStereoDataset('training_data.xlsx', 3, 15, True)
# getDifference('training_data.xlsx', 3, 4, 9, 25, 7, 12, 26)
# generatePairingDataset('training_data.xlsx', 5, 19, reducePair=True)
# removeBus('training_data.xlsx', 3, 5)
#generatePairingDataset('training_data.xlsx', 6, 19)
# generateStereoDataset('training_data.xlsx', 4, 15)#, True)
# removeBus('training_data.xlsx', 4, 6)
# getDifference('training_data.xlsx', 4, 4, 9, 25, 7, 12, 26)
#normalizeDistance('training_data.xlsx', 7, 16, 2464)
#verticalDistance('training_data.xlsx', 5, 6, 9, 20, 1632)
# generatePairingDataset('training_data - Copy.xlsx', 6, 13)
#removeBus('training_data - Copy.xlsx', 4, 7)
# removeBus('training_data.xlsx', stereoIdx=4)
# #extractVidToFrame('MVI_1310.MOV')
# reduceCell('training_data.xlsx', 5, 13)
# createRatioColumn('training_data.xlsx')
#reduceCell('training_data - Copy.xlsx', 5, 13)
# createRangeColumn('training_data.xlsx')
# createRatioColumn('training_data.xlsx')
#copyCell('training_data - Copy.xlsx', 4, 5)

#createRangeColumn('training_data (3).xlsx')
# # # extractVidToFrame('MVI_1310.MOV')
# # #resize('Img_2\\left_origin\\', 'Img_2\\left\\')
# # #show('Img_2\\left\\25.JPG', 'Img_2\\right\\25.JPG')
# # # sift = cv2.xfeatures2d.SIFT_create()
# # # img = cv2.imread('LED input.png')
# # # gray= cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
# # # kps, des = sift.detectAndCompute(gray, None)
# # # # kps = [cv2.KeyPoint(i, j, 1) for j in range(11, 20) for i in range(0, 10)]
# # # # for kp in kps:
# # # #     print(kp.pt)

# # # #print(np.shape(des[0, :]))
# # # print(des[0, :])

# # import matplotlib.pyplot as plt
# # import matplotlib as mpl
# # import numpy as np

# # # Make an example plot with two subplots...
# # fig = plt.figure()
# # ax1 = fig.add_subplot(2,1,1)
# # ax1.plot(range(10), 'b-')

# # ax2 = fig.add_subplot(2,1,2)
# # ax2.plot(range(20), 'r^')

# # plt.show()

# # # Save the full figure...
# # fig.savefig('full_figure.png')

# # # Save just the portion _inside_ the second axis's boundaries
# # extent = ax2.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
# # fig.savefig('ax2_figure.png', bbox_inches=extent)

# # # Pad the saved area by 10% in the x-direction and 20% in the y-direction
# # fig.savefig('ax2_figure_expanded.png', bbox_inches=extent.expanded(1.1, 1.2))

# import cv2
# import numpy as np
# from matplotlib import pyplot as plt
# import timeit
# #import skimage

# img = cv2.imread('template.png', 0)
# img2 = img.copy()
# template = cv2.imread('template.png', 0)
# #template = cv2.flip(template, 1)
# w, h = template.shape[::-1]

# # All the 6 methods for comparison in a list
# methods = ['cv2.TM_CCOEFF', 'cv2.TM_CCOEFF_NORMED', 'cv2.TM_CCORR',
#             'cv2.TM_CCORR_NORMED', 'cv2.TM_SQDIFF', 'cv2.TM_SQDIFF_NORMED']

# f = open('result_perfect_3.txt', 'w+')

# for meth in methods:
#     img = img2.copy()
#     method = eval(meth)

#     # Apply template Matching
#     start = timeit.default_timer()
#     res = cv2.matchTemplate(img, template, method)
#     print(res)
#     min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
#     flat = res.flatten()
#     flat.sort()
#     # second_min_val = flat[1]
#     # second_max_val = flat[-2]
#     stop = timeit.default_timer()

#     # If the method is TM_SQDIFF or TM_SQDIFF_NORMED, take minimum
#     if method in [cv2.TM_SQDIFF, cv2.TM_SQDIFF_NORMED]:
#         top_left = min_loc
#         score = min_val
#         #second_score = second_min_val
#     else:
#         top_left = max_loc
#         score = max_val
#         #second_score = second_max_val
#     # ratio = abs(score - second_score) / float(second_score)
#     bottom_right = (top_left[0] + w, top_left[1] + h)
#     cv2.rectangle(img, top_left, bottom_right, 255, 4)
#     #f.write(meth + ':' + str(top_left) + ',' + str(bottom_right) + ':' + str(score) + '(' + str(stop - start) + ')' + '\n')
#     f.write(meth + '. Position: (' + str(top_left) + ',' + str(bottom_right) + '). First score: ' + str(score) + '\n')# + '. Second score: ' + str(second_score) + '. Ratio: ' + str(ratio) + '. Time: ' + str(stop - start) + ' s\n')
#     plt.subplot(221)
#     plt.imshow(img2, cmap = 'gray')
#     plt.title('Original image') 
#     plt.xticks([]), plt.yticks([])
#     plt.subplot(222)
#     plt.imshow(template, cmap = 'gray')
#     plt.title('Template') 
#     plt.xticks([]), plt.yticks([])
#     plt.subplot(223)
#     plt.imshow(res, cmap = 'gray')
#     plt.title('Matching Result') 
#     plt.xticks([]), plt.yticks([])
#     plt.subplot(224)
#     plt.imshow(img, cmap = 'gray')
#     plt.title('Detected Point')
#     plt.xticks([]), plt.yticks([])
#     plt.suptitle(meth)

#     plt.show()

# f.close()
#import os
# import numpy as np
# import pickle
# def read_from_pickle(path):
#     print('Access')
#     with open(path, 'rb') as file:
#         try:
#             print('Hi')
#             while True:
#                 print('Good')
#                 yield pickle.load(file)
#         except EOFError:
#             print('Hello')
#             pass

# read_from_pickle('mnist.pkl')

# print('Access')
# #objects = []
# #readFile = open('output.txt', 'w+')
# with open('mnist.pkl', 'rb') as file:
#     try:
#         #print('Hi')
#         # file = pickle._Unpickler(file)
#         # file.encoding = 'latin1'
#         while True:
#             print('Good')
#             #_object = file.load()
#             _object = pickle.load(file, encoding='latin1')
#             #objects.append(_object)
#             print(str(type(_object[0][1][0])) + '\n')
#     except EOFError:
#         print('Hello')
#         #pass
# print(np.inf)
#readFile.close()

# from openpyxl import load_workbook
'''