import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
import sys
import utils
import ntpath
from openpyxl import load_workbook


# from skimage.feature import haar_like_feature_coord
# from skimage.feature import draw_haar_like_feature
from Region import CList

# utils.min_H_1 = 342 / 2
# utils.max_H_1 = 180
# utils.min_H_2 = 0
# utils.max_H_2 = 9 / 2
# utils.max_S = 255
# utils.min_S = 0.4645 * utils.max_S
# utils.max_V = 255
# utils.min_V = 0.2 * utils.max_S

#training_key = {'Image ID': 1, 'IDL': 2, 'IDR': 3, 'DBL': 4, 'WL': 5, 'HL': 6, 'VL': 7, 'DNL': 8, 'DBR': 9,	'WR': 10, 'HR': 11, 'VR': 12, 'DNR': 13, 'NCC': 14, 'M': 15, 'SIFT': 16, 'SURF': 17, 'ORB': 18, 'DNNL': 19, 'DNNR': 20, 'SL': 21, 'SR': 22, 'Vertical ratio': 23, 'Size ratio': 24}

# pairing_key = {
# 'Image ID': 1, 
# 'IDL': 2, 
# 'IDR': 3, 
# 'WL': 4, 
# 'HL': 5, 
# 'VL': 6, 
# 'WR': 7, 
# 'HR': 8, 
# 'VR': 9, 
# 'NCC': 10, 
# 'SIFT': 11, 
# 'SURF': 12, 
# 'ORB': 13, 
# 'SL': 14, 
# 'SR': 15, 
# 'Inter-distance': 16, 
# 'Vertical ratio': 17, 
# 'Size ratio': 18, 
# 'M': 19}

def thresholdColor(im):
    '''Apply color thresholding on input image
    @im [in]: input image
    @thres [out]: thresholding result
    '''
    im_hsv = cv2.cvtColor(im, cv2.COLOR_BGR2HSV)
    
    #mask with hue range 2: from 0 degree to 9 degree
    im_red_ball_mask_1 = cv2.inRange(im_hsv, (utils.min_H_1, utils.min_S, utils.min_V), (utils.max_H_1, utils.max_S, utils.max_V))
    im_red_ball_mask_1 = im_red_ball_mask_1.astype('bool')
    
    #mask with hue range 2: from 342 degree to 360 degree
    im_red_ball_mask_2 = cv2.inRange(im_hsv, (utils.min_H_2, utils.min_S, utils.min_V), (utils.max_H_2, utils.max_S, utils.max_V)) 
    im_red_ball_mask_2 = im_red_ball_mask_2.astype('bool')

    # combine 2 masks to apply on image
    im_red_ball_mask_full = im_red_ball_mask_1 + im_red_ball_mask_2

    #color threshold result
    thres = im * np.dstack((im_red_ball_mask_full, im_red_ball_mask_full, im_red_ball_mask_full))
    return thres

def closeTransform(im):
    '''Apply morphological close transform on input image
    @im [in]: input image
    @closing [out]: morphological close result
    '''
    kernel = np.ones((12, 12), np.uint8) 
    closing = cv2.morphologyEx(im, cv2.MORPH_CLOSE, kernel)
    return closing    

def isForeground(img, x_axis, y_axis):
    ''' Check if the point in img with coord (x_axis, y_axis) is the foreground point or not
    @img [in]: input image
    @x_axis [in]: x-axis coordinate of point
    @y_axis [in]: y-axis coordinate of point
    @isForeground [out]: true if the point is foreground and vice versa 
    '''
    isForeground = img[x_axis, y_axis, 0] != 0 or img[x_axis, y_axis, 1] != 0 or img[x_axis, y_axis, 2] != 0
    return isForeground

def isBoundary(img, x_axis, y_axis):
    ''' Determine if the point in img with coord (x_axis, y_axis) is on image boundary or not based on checking its neighboorhood
    @img [in]: input image
    @x_axis [in]: x-axis coordinate of point
    @y_axis [in]: y-axis coordinate of point
    @isBoundary: true if the point is on boundary and vice versa
    '''
    width, height,_ = img.shape
    if 0 < x_axis < width and 0 < y_axis < height:
        num_of_neighbor = 8
    elif (0 < x_axis < width and  (y_axis == 0 or  y_axis == height - 1)) or (0 < y_axis < height and  (x_axis == 0 or  x_axis == width - 1)):
        num_of_neighbor = 5
    else:
        num_of_neighbor = 3
    num_of_foreground = isForeground(img, x_axis - 1, y_axis - 1) + isForeground(img, x_axis - 1, y_axis) + isForeground(img, x_axis - 1, y_axis + 1)\
        + isForeground(img, x_axis, y_axis - 1) + isForeground(img, x_axis, y_axis + 1)\
        + isForeground(img, x_axis + 1, y_axis - 1) + isForeground(img, x_axis + 1, y_axis) +  isForeground(img, x_axis + 1, y_axis + 1)
    is_Boundary = (0 < num_of_foreground < num_of_neighbor)
    return is_Boundary

def assignLabel(img, origin, filename=None, excelFile=None):
    '''Assign region label to the image
    @img [in]: labelled image
    @origin [in]: origin image
    @filename [out]: output file name
    @excelFile [in]: excel file to record data
    '''
    width, height, _ = img.shape
    cur_label = 1
    r_List = CList(img, origin)#width, height)
    for i in range(width):
        for j in range(height):
            if isForeground(img, i, j):# is True:
                '''Get minimum label from west & north neighbor'''
                #label_00 = rg.CList.mGetInstance(width, height).mGetLabel(i - 1, j - 1)#consider img[i - 1, j - 1]
                #label_01 = rg.CList.mGetInstance(width, height).mGetLabel(i - 1, j)#consider img[i - 1, j]
                #label_02 = rg.CList.mGetInstance(width, height).mGetLabel(i - 1, j + 1)#consider img[i - 1, j + 1]
                #label_10 = rg.CList.mGetInstance(width, height).mGetLabel(i, j - 1)#consider img[i, j - 1]
                #min_label = np.amin([label_00, label_01, label_02, label_10])
                label_00 = r_List.mGetLabel(i - 1, j - 1)
                label_01 = r_List.mGetLabel(i - 1, j)
                label_02 = r_List.mGetLabel(i - 1, j + 1)
                label_10 = r_List.mGetLabel(i, j - 1)
                min_label = np.amin([label_00, label_01, label_02, label_10])
                #min_label = np.amin([label_01, label_10])
                '''Assign label for current point'''
                '''@1st case: exist minimum label from precede neighbors, assign this label to the current point, along
                with merge the region of each of neighbor to this current point's neighbor
                @2nd case: minimum label from precedent neighbor not exist (both north & west neighbors are background), assign 
                new label (new region) for the current point
                '''
                if min_label != np.inf:
                    if label_00 > min_label:
                        r_List.mMergeRegions(label_00, min_label)
                    if label_01 > min_label:
                        r_List.mMergeRegions(label_01, min_label)
                    if label_02 > min_label:
                        r_List.mMergeRegions(label_02, min_label)
                    if label_10 > min_label:
                        r_List.mMergeRegions(label_10, min_label)
                    # rg.CList.mGetInstance(width, height).mMergeRegions(label_01, min_label)  
                    # rg.CList.mGetInstance(width, height).mMergeRegions(label_10, min_label) 
                    # rg.CList.mGetInstance(width, height).mAppend(min_label, i, j, img[i, j])
                    # r_List.mMergeRegions(label_01, min_label)
                    # r_List.mMergeRegions(label_10, min_label)
                    r_List.mAppend(min_label, i, j, img[i, j])
                else:
                    # rg.CList.mGetInstance(width, height).mAppend(cur_label, i, j, img[i, j])
                    r_List.mAppend(cur_label, i, j, img[i, j])
                    cur_label += 1
    # r_List.mSuppress()
    # if filename is not None and excelFile is not None:
    #     r_List.mWriteToFile(filename, excelFile)
    return r_List

def showResult(im_title, im1, im2, rList):#:im3, vertices):#, kps):
    ''' This is a function to display the result
    @im_title [in]: image title
    @im1 [in]: original image
    @im2 [in]: labelled image
    @rList [in]: region list
    '''
    im3 = rList.mGetLabelMat()
    vertices, _ = rList.mGetBoundaries()
    plt.figure()  
    plt.title(im_title)
    plt.axis("off")
    ax1 = plt.subplot(2, 2, 1)
    ax2 = plt.subplot(2, 2, 2, sharex= ax1, sharey=ax1)
    ax3 = plt.subplot(2, 2, 4, sharex= ax1, sharey=ax1)
    ax4 = plt.subplot(2, 2, 3, sharex= ax1, sharey=ax1)
    if len(im1.shape) == 2:
        ax1.imshow(im1, cmap="gray")
    else:
        im1_display = cv2.cvtColor(im1, cv2.COLOR_RGB2BGR)
        ax1.imshow(im1_display)
    if len(im2.shape) == 2:
        ax2.imshow(im2, cmap="gray")
        ax3.imshow(im2, cmap="gray")
    else:
        im2_display = cv2.cvtColor(im2, cv2.COLOR_RGB2BGR)
        ax2.imshow(im2_display)
        im3_display = cv2.cvtColor(im1, cv2.COLOR_RGB2BGR)
                
        # sift = cv2.xfeatures2d.SIFT_create()
        # kp = sift.detect(im3_display, None)
        # pts = np.asarray([kp[idx].pt for idx in range(0, len(kp))])#.reshape(-1, 1, 2)
        # file = open("keypoint.txt", "w+")
        # for pt in pts:
        #     file.write(str(int(im3[int(pt[1]), int(pt[0])])) + ':(' + str(int(pt[0])) + ',' + str(int(pt[1])) + ')\n')
        # file.close()
        # im3_display = cv2.drawKeypoints(im3_display, kp, im3_display)#, flags=cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)
        # hsv = cv2.cvtColor(im3_display, cv2.COLOR_BGR2HSV)
        # try:
        #     h, s, v = cv2.split(hsv)
        #     # kps = []
        #     # des = np.array([])
        #     print('I am in\n')
        #     sift = cv2.xfeatures2d.SIFT_create()
        #     kp_h = sift.detect(h, None)
        #     #kps.append(kp_h)
        #     np.append(des, des_h)
        #     kp_s = sift.detect(s, None)
        #     #kps.append(kp_s)
        #     np.append(des, des_s)
        #     kp_v = sift.detect(v, None)
        #     #kps.append(kp_v)
        #     #np.append(des, des_v)  
        #     im3_display = cv2.drawKeypoints(im3_display, kp_h, im3_display)
        #     im3_display = cv2.drawKeypoints(im3_display, kp_s, im3_display)
        #     im3_display = cv2.drawKeypoints(im3_display, kp_v, im3_display)
        #     #return kps, des
        # except:
        #     print('I am out\n')
            #return None
        # for x in kps:
        #     kp = kps.get(x)
        #     if kp is not None and kp != []:
        #         print(kp.pt[0])
        #         print('---')
        #         try:
        #             im3_display = cv2.drawKeypoints(im3_display, kp, im3_display)
        #         except: #expression as identifier:
        #             print('Cannot draw')
        #     else:
        #         print('Keypoint is none')
        color = (0, 255, 0) 
        thickness = 3
        index = 0
        for row in vertices:
            if row[1] != row[3] and row[0] != row[2]:
                cv2.rectangle(im3_display, (row[1], row[0]), (row[3], row[2]), color, thickness)
                #cv2.putText(im3_display, str(label[index]), (row[1], row[0] - 10), cv2.FONT_HERSHEY_SIMPLEX, 1.5, color, thickness)
            index += 1
        ax3.imshow(im3_display)
        cv2.imwrite('Output\\Img\\OutputRect_11_left.jpg', cv2.cvtColor(im3_display, cv2.COLOR_RGB2BGR))
    ax4.imshow(im3, cmap="gray")
    cv2.imwrite('Output\\Img\\OutputRect_grayscale.jpg', im3)
    plt.show()

def saveOutput(im, folder_name, rList):
    ''' This is a function to save the output
    @im [in]: image title
    @folder_name [in]: folder path 
    @rList [in]: region list
    '''
    im3_display = cv2.cvtColor(im, cv2.COLOR_RGB2BGR)
    im2_display = cv2.cvtColor(im, cv2.COLOR_RGB2BGR)
    vertices, label = rList.mGetBoundaries()
    color = (0, 255, 0) 
    thickness = 1
    index = 0
    for row in vertices:
        if row[1] != row[3] and row[0] != row[2]:
            cv2.rectangle(im3_display, (row[1], row[0]), (row[3], row[2]), color, thickness)
            cv2.rectangle(im2_display, (row[1], row[0]), (row[3], row[2]), color, thickness)
            cv2.putText(im3_display, str(label[index]), (row[1], row[0] - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, color, thickness)
        index += 1
    folder_name = folder_name[folder_name.find('\\'): ]
    child_file = 'Output' + folder_name
    child_file_2 = 'Output_2' + folder_name
    child_folder = os.path.splitext(child_file)[0]
    child_folder = child_folder[: child_folder.find('\\', -1) - 1]
    child_folder_2 = os.path.splitext(child_file_2)[0]
    child_folder_2 = child_folder_2[: child_folder_2.find('\\', -1) - 1]
    # print(child_file)
    # print(child_folder)
    # print(child_file_2)
    # print(child_folder_2)
    if  os.path.isdir(child_folder) is False:
        os.makedirs(child_folder)
    #     #print('Accept child_folder!')
    if  os.path.isdir(child_folder_2) is False:
        os.makedirs(child_folder_2)
        #print('Accept child_folder_2!')
    cv2.imwrite(child_file, cv2.cvtColor(im3_display, cv2.COLOR_RGB2BGR))
    cv2.imwrite(child_file_2, cv2.cvtColor(im2_display, cv2.COLOR_RGB2BGR))
    #cv2.imwrite()

def detectLEDRegion(filename, desname=None, excelFile=None):
    '''Entire process of detect LED region framework
    @filename [in]: image file name
    @desname [in]: output file name
    @excelFile [in]: excel file
    '''
    #print('Hello')
    im = cv2.imread(filename)
    #print('Shape: ', np.shape(im))
    thres = thresholdColor(im)
    #closing = closeTransform(thres)
    #print('Hi')
    #rList = assignLabel(closing, desname, excelFile)
    rList = assignLabel(thres, im)#, desname, excelFile)
    #print('Bar')
    rList.mSuppress()
    if filename is not None:# and excelFile is not None:
        rList.mWriteToFile(filename, excelFile)
    #showResult("LED region detection result", im, closing, rList)
    saveOutput(im, filename, rList)
    return rList
    # rList.mGetPair(utils.FD.DENSE_ORB)

def detectLEDRegion_2(im):
    '''Entire process of detect LED region framework (backup)
    @im [in]: input image
    '''
    thres = thresholdColor(im)
    closing = closeTransform(thres)
    return closing, assignLabel(closing)

def showPair(imgs):
    '''Show output images by pair
    @imgs [in]: input image folder
    '''
    left_img = cv2.imread('Img_2\\left\\11.JPG')
    gray_left = cv2.cvtColor(left_img, cv2.COLOR_BGR2GRAY)
    width, height, _ = left_img.shape
    right_img = cv2.imread('Img_2\\right\\11.JPG')
    gray_right = cv2.cvtColor(right_img, cv2.COLOR_BGR2GRAY)
    left_res = cv2.imread('Img_2\\left\\11.JPG')
    right_res = cv2.imread('Img_2\\right\\11.JPG')
    #left_img = cv2.cvtColor(imgs[2][0], cv2.COLOR_RGB2BGR)
    #right_img = cv2.cvtColor(imgs[2][1], cv2.COLOR_RGB2BGR)
    # left_res = cv2.cvtColor(imgs[2][0], cv2.COLOR_RGB2BGR)
    # right_res = cv2.cvtColor(imgs[2][1], cv2.COLOR_RGB2BGR)
    color = (0, 255, 0) 
    thickness = 2
    left_clos, left_rList = detectLEDRegion_2(left_res)  
    right_clos, right_rList = detectLEDRegion_2(right_res) 
    left_img = cv2.cvtColor(left_img, cv2.COLOR_RGB2BGR)
    right_img = cv2.cvtColor(right_img, cv2.COLOR_RGB2BGR)
    left_res = cv2.cvtColor(left_res, cv2.COLOR_RGB2BGR)
    right_res = cv2.cvtColor(right_res, cv2.COLOR_RGB2BGR)
    left_ver, _ = left_rList.mGetBoundaries()
    right_ver, _ = right_rList.mGetBoundaries()
    # left_kp = [cv2.KeyPoint(x, y, 1) for y in range(0, height) 
    #                                 for x in range(0, width)]
    # right_kp = left_kp
    #sift = cv2.xfeatures2d.SIFT_create()    
    # left_des, left_kp = left_rList.mGetListKeypoints(gray_left)
    # right_des, right_kp = right_rList.mGetListKeypoints(gray_right)
    # # left_des = sift.compute(left_clos, left_kp)
    # # right_des = sift.compute(right_clos, right_kp)
    # # right_des, right_kp = right_rList.mGetListKeypoints(right_clos)
    # # create BFMatcher object
    # bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    # # Match descriptors.
    # matches = bf.match(left_des, right_des)
    # # Sort them in the order of their distance.
    # matches = sorted(matches, key = lambda x:x.distance)
    for row in left_ver:
        if row[1] != row[3] and row[0] != row[2]:
            cv2.rectangle(left_res, (row[1], row[0]), (row[3], row[2]), color, thickness)
    for row in right_ver:
        if row[1] != row[3] and row[0] != row[2]:
            cv2.rectangle(right_res, (row[1], row[0]), (row[3], row[2]), color, thickness)        
    plt.figure() 
    plt.axis("off")
    ax1 = plt.subplot(2, 2, 1)
    ax2 = plt.subplot(2, 2, 2, sharex=ax1, sharey=ax1)
    ax3 = plt.subplot(2, 2, 3, sharex=ax1, sharey=ax1)
    ax4 = plt.subplot(2, 2, 4, sharex=ax1, sharey=ax1)
    ax1.imshow(left_img)#, cmap=cv2.COLOR_RGB2BGR)
    ax2.imshow(right_img)#, cmap=cv2.COLOR_RGB2BGR)
    ax3.imshow(left_res)#, cmap=cv2.COLOR_RGB2BGR)
    ax4.imshow(right_res)#, cmap=cv2.COLOR_RGB2BGR)
    cv2.imwrite('OutputRect_left.jpg', cv2.cvtColor(left_res, cv2.COLOR_RGB2BGR))
    cv2.imwrite('OutputRect_right.jpg', cv2.cvtColor(right_res, cv2.COLOR_RGB2BGR))
    plt.show()
    # img3 = cv2.drawMatches(left_img, left_kp, right_img, right_kp, matches, flags=2)
    # plt.imshow(img3)
    # plt.show()

def loadImages(folder_name):
    '''Load images from input folder
    @folder_name [in]: input folder name
    @pair_list [out]: list of images by left-right pair (exception return None)
    '''
    left_dir = folder_name + '\\left_bak'
    right_dir = folder_name + '\\right'
    left_dir_out = folder_name + '\\left_bak_out'
    right_dir_out = folder_name + '\\right_out'
    n = len(next(os.walk(left_dir))[2])
    #print(next(os.walk(left_dir))[2])
    #print(left_dir)
    #pair_list = []
    try:
        for r, _, f in os.walk(left_dir):
            for file in f:
                if '.JPG' in file:
                    #print('shape: ' + np.shape(file))
                    imgDir = os.path.join(r, file)
                    outDir = left_dir_out + '\\' + os.path.splitext(file)[0] + '.txt'
                    #print(imgDir + ': ' + outDir + '\n')
                    detectLEDRegion(imgDir, outDir)
        for r, _, f in os.walk(right_dir):
            for file in f:
                if '.JPG' in file:
                    imgDir = os.path.join(r, file)
                    outDir = right_dir_out + '\\' + os.path.splitext(file)[0] + '.txt'
                    #print(imgDir + ': ' + outDir + '\n')
                    detectLEDRegion(imgDir, outDir)
    except: #expression as identifier:
        print("Unexpected error:", sys.exc_info()[0])

def generateDataset(excelFile, imgUrl, matching=2, pairingLeft=4, pairingRight=6):
    '''Generate dataset from images
    @excelFile [in]: dataset excel file
    @imgUrl [in]: folder contains input images
    @matching [in]: sheet for stereo matching
    @pairingLeft [in]: sheet for pairing in left folder
    @pairingRight [in]: sheet for pairing in right folder
    '''
    wb = load_workbook(excelFile)
    sheet = wb.worksheets[matching]
    sheet_2 = wb.worksheets[pairingLeft]
    sheet_3 = wb.worksheets[pairingRight]
    left_dir = imgUrl + '\\left'
    left_dir_out = left_dir + '_out'
    right_dir = imgUrl + '\\right'
    right_dir_out = right_dir + '_out'
    if os.path.isdir(left_dir_out) is False:
        os.makedirs(left_dir_out)
    if os.path.isdir(right_dir_out) is False:
        os.makedirs(right_dir_out)
    imgList = next(os.walk(left_dir))[2]
    idx = 2
    idx_2 = 2
    idx_3 = 2
    for name in imgList:
        lImg = left_dir + '\\' + name
        lOutDir = left_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        lList = detectLEDRegion(lImg, lOutDir, excelFile)
        rImg = right_dir + '\\' + name
        rOutDir = right_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        rList = detectLEDRegion(rImg, rOutDir, excelFile)
        lListId = lList.mGetLabelList()
        rListId = rList.mGetLabelList()

        lLen = len(lListId)

        '''Taillight pairing in left folder'''
        for i in range(lLen - 1):
            _, lw, lh, lv, _, _, ls = lList.mGetRegionInfo(lListId[i], utils.STEREO.LEFT)
            for j in range(i + 1, lLen):
                _, lw_2, lh_2, lv_2, _, _, ls_2 = lList.mGetRegionInfo(lListId[j], utils.STEREO.LEFT)
                inter_distance = lList.mGetInterDistance(lListId[i], lListId[j])
                ncc_2 = lList.mGetNCCValue(lListId[i], lList, lListId[j], True)
                sift_2 = 0#lList.mGetDescriptionDiff(lListId[i], lList, lListId[j], utils.FD.DENSE_SIFT, True) 
                surf_2 = 0
                #surf_2 = lList.mGetDescriptionDiff(lListId[i], lList, lListId[j], utils.FD.DENSE_SURF, True)   #mGetNCCValue(lListId[i], lList, lListId[j], True)
                orb_2 = 0#lList.mGetDescriptionDiff(lListId[i], lList, lListId[j], utils.FD.DENSE_ORB, True)
                vr = float(lv) / float(lv_2)
                sr = float(ls) / float(ls_2)
                features_2 = [name, lListId[i], lListId[j], lw, lh, lv, lw_2, lh_2, lv_2, ncc_2, sift_2, surf_2, orb_2, ls, ls_2, inter_distance, vr, sr]
                for k in range(18):
                    # if features_2[k] == np.inf:
                    #     features_2[k] = 1
                    sheet_2.cell(row=idx_2, column=k + 1).value = features_2[k]
                # print('SIFT 2: ', sift_2)
                # sheet_2.cell(row=idx_2, column=utils.training_key['SIFT']).value = sift_2
                # print('Cell 2 SIFT: ', sheet_2.cell(row=idx_2, column=utils.training_key['SIFT']).value)
                # #sheet_2.cell(row=idx_2, column=utils.training_key['SURF']).value = surf_2
                # sheet_2.cell(row=idx_2, column=utils.training_key['ORB']).value = orb_2
                # print('Cell 2 ORB: ', sheet_2.cell(row=idx_2, column=utils.training_key['ORB']).value)
                idx_2 += 1  

        rLen = len(rListId)
        '''Taillight pairing in right folder'''
        for i in range(rLen - 1):
            _, rw, rh, rv, _, _, rs = rList.mGetRegionInfo(rListId[i], utils.STEREO.RIGHT)
            for j in range(i + 1, rLen):
                _, rw_3, rh_3, rv_3, _, _, rs_3 = rList.mGetRegionInfo(rListId[j], utils.STEREO.RIGHT)
                inter_distance = rList.mGetInterDistance(rListId[i], rListId[j])
                ncc_3 = rList.mGetNCCValue(rListId[i], rList, rListId[j], True)
                sift_3 = 0#rList.mGetDescriptionDiff(rListId[i], rList, rListId[j], utils.FD.DENSE_SIFT, True) 
                surf_3 = 0
                #surf_3 = rList.mGetDescriptionDiff(rListId[i], rList, rListId[j], utils.FD.DENSE_SURF, True)
                orb_3 = 0#rList.mGetDescriptionDiff(rListId[i], rList, rListId[j], utils.FD.DENSE_ORB, True) 
                vr = float(lv) / float(lv_2)
                sr = float(ls) / float(ls_2)
                features_3 = [name, rListId[i], rListId[j], rw, rh, rv, rw_3, rh_3, rv_3, ncc_3, sift_3, surf_3, orb_3, rs, rs_3, inter_distance, vr, sr]
                for k in range(18):
                    # if features_2[k] == np.inf:
                    #     features_2[k] = 1
                    sheet_3.cell(row=idx_3, column=k + 1).value = features_3[k]
                # print('SIFT 3: ', sift_3)
                # sheet_3.cell(row=idx_3, column=utils.training_key['SIFT']).value = sift_3
                # #sheet_3.cell(row=idx_3, column=utils.training_key['SURF']).value = surf_3
                # sheet_3.cell(row=idx_3, column=utils.training_key['ORB']).value = orb_3
                # print('Cell 3 SIFT: ', sheet_3.cell(row=idx_3, column=utils.training_key['SIFT']).value)
                # print('Cell 3 ORB: ', sheet_3.cell(row=idx_3, column=utils.training_key['ORB']).value)
                idx_3 += 1  
        
        '''utils.STEREO taillight matching'''        
        for lid in lListId:
            lbd, lw, lh, lv, ldn, ldnn, ls = lList.mGetRegionInfo(lid, utils.STEREO.LEFT)
            for rid in rListId:
                rbd, rw, rh, rv, rdn, rdnn, rs = rList.mGetRegionInfo(rid, utils.STEREO.RIGHT)
                ncc = lList.mGetNCCValue(lid, rList, rid)
                sift = 0#lList.mGetDescriptionDiff(lid, rList, rid, utils.FD.DENSE_SIFT)#, False) 
                #surf = lList.mGetDescriptionDiff(lid, rList, rid, utils.FD.DENSE_SURF, True)
                orb = 0#lList.mGetDescriptionDiff(lid, rList, rid, utils.FD.DENSE_ORB)#, False)
                features = [name, lid, rid, lbd, lw, lh, lv, ldn, rbd, rw, rh, rv, rdn, ncc]
                for i in range(14):
                    sheet.cell(row=idx, column=i + 1).value = features[i]
                #print('SIFT 1: ', sift)
                sheet.cell(row=idx, column=utils.training_key['SIFT']).value = sift
                #sheet.cell(row=idx, column=utils.training_key['SURF']).value = surf
                sheet.cell(row=idx, column=utils.training_key['ORB']).value = orb
                sheet.cell(row=idx, column=utils.training_key['DNNL']).value = ldnn
                sheet.cell(row=idx, column=utils.training_key['DNNR']).value = rdnn
                sheet.cell(row=idx, column=utils.training_key['SL']).value = ls
                sheet.cell(row=idx, column=utils.training_key['SR']).value = rs
                sheet.cell(row=idx, column=utils.training_key['Size ratio']).value = float(ls) / float(rs)
                sheet.cell(row=idx, column=utils.training_key['Vertical ratio']).value = float(lv) / float(rv)
                # print('Cell SIFT: ', sheet.cell(row=idx, column=utils.training_key['SIFT']).value)
                # print('Cell ORB: ', sheet.cell(row=idx, column=utils.training_key['ORB']).value)
                idx = idx + 1
                print(lid, rid)

    wb.save(excelFile)         

def generateMatchingDataset(excelFile, imgUrl):
    '''Generate matching dataset:
    @excelFile [in]: dataset excel file,
    @imgUrl [in]: folder contains input images
    '''
    wb = load_workbook(excelFile)
    sheet = wb.worksheets[0]
    left_dir = imgUrl + '\\left'
    left_dir_out = left_dir + '_out'
    right_dir = imgUrl + '\\right'
    right_dir_out = right_dir + '_out'
    if os.path.isdir(left_dir_out) is False:
        os.makedirs(left_dir_out)
    if os.path.isdir(right_dir_out) is False:
        os.makedirs(right_dir_out)
    imgList = next(os.walk(left_dir))[2]
    idx = 2
    for name in imgList:
        lImg = left_dir + '\\' + name
        lOutDir = left_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        lList = detectLEDRegion(lImg, lOutDir, excelFile)
        rImg = right_dir + '\\' + name
        rOutDir = right_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        rList = detectLEDRegion(rImg)#, rOutDir, excelFile)
        lListId = lList.mGetLabelList()
        rListId = rList.mGetLabelList()
        for lid in lListId:
            lbd, lw, lh, lv, ldn, _, ls = lList.mGetRegionInfo(lid, utils.STEREO.LEFT)
            for rid in rListId:
                rbd, rw, rh, rv, rdn, _, rs = rList.mGetRegionInfo(rid, utils.STEREO.RIGHT)
                ncc, ncc_2 = lList.mGetNCCValue(lid, rList, rid)
                bhatta = lList.mGetBhattaValue(lid, rList, rid)
                wr = min(lw, rw) / max(lw, rw)
                if lw < rw:
                    wr = -wr
                hr = min(lh, rh) / max(lh, rh)
                if lh < rh:
                    hr = -hr
                dbr = min(lbd, rbd) / max(lbd, rbd)
                if dbr < dbr:
                    dbr = -dbr    
                vr = min(lv, rv) / max(lv, rv)
                if lv < rv:
                    vr = -vr 
                sr = min(ls, rs) / max(ls, rs)
                if ls < rs:
                    sr = -sr               
                # sift = 0
                # orb = 0
                features = [name, lid, rid, lbd, lw, lh, lv, ldn, rbd, rw, rh, rv, rdn, ncc, ncc_2, bhatta, wr, hr, dbr, vr, sr]
                features_len = len(features)
                for i in range(features_len):
                    sheet.cell(row=idx, column=i + 1).value = features[i]
                # sheet.cell(row=idx, column=utils.training_key['SIFT']).value = sift
                # sheet.cell(row=idx, column=utils.training_key['ORB']).value = orb
                # sheet.cell(row=idx, column=utils.training_key['DNNL']).value = ldnn
                # sheet.cell(row=idx, column=utils.training_key['DNNR']).value = rdnn
                # sheet.cell(row=idx, column=utils.training_key['SL']).value = ls
                # sheet.cell(row=idx, column=utils.training_key['SR']).value = rs
                # sheet.cell(row=idx, column=utils.training_key['Size ratio']).value = float(ls) / float(rs)
                # sheet.cell(row=idx, column=utils.training_key['Vertical ratio']).value = float(lv) / float(rv)
                idx = idx + 1
                print(lid, rid)
    wb.save(excelFile)

def generatePairingDataset(excelFile, imgUrl):
    '''Generate pairing dataset:
    @excelFile [in]: dataset excel file,
    @imgUrl [in]: folder contains input images
    '''
    wb = load_workbook(excelFile)
    sheet = wb.worksheets[0]
    dir_out = imgUrl + '_out'
    if os.path.isdir(dir_out) is False:
        os.makedirs(dir_out)
    imgList = next(os.walk(imgUrl))[2]
    idx = 2
    for name in imgList:
        lImg = imgUrl + '\\' + name
        lOutDir = dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        lList = detectLEDRegion(lImg)#, lOutDir, excelFile)
        lListId = lList.mGetLabelList()
        lLen = len(lListId)
        for i in range(lLen - 1):
            _, lw, lh, lv, _, _, ls = lList.mGetRegionInfo(lListId[i], utils.STEREO.LEFT)
            for j in range(i + 1, lLen):
                _, lw_2, lh_2, lv_2, _, _, ls_2 = lList.mGetRegionInfo(lListId[j], utils.STEREO.LEFT)
                inter_distance = lList.mGetInterDistance(lListId[i], lListId[j])
                ncc, ncc_2 = lList.mGetNCCValue(lListId[i], lList, lListId[j], True)
                bhatta = lList.mGetBhattaValue(lListId[i], lList, lListId[j], True)
                wr = min(lw, lw_2) / max(lw, lw_2)
                if lw < rw:
                    wr = -wr
                hr = min(lh, lh_2) / max(lh, lh_2)
                if lh < rh:
                    hr = -hr
                vr = min(lv, lv_2) / max(lv, lv_2)
                if lv < rv:
                    vr = -vr 
                sr = min(ls, ls_2) / max(ls, ls_2)
                if ls < rs:
                    sr = -sr
                vdr = lList.mGetVerticalCentroidRatio(lListId[i], lListId[j])
                # sift_2 = 0
                # surf_2 = 0
                # orb_2 = 0
                # vr = float(lv) / float(lv_2)
                # sr = float(ls) / float(ls_2)
                features_2 = [name, lListId[i], lListId[j], lw, lh, lv, lw_2, lh_2, lv_2, inter_distance, ncc, ncc_2, bhatta, wr, hr, vr, sr, vdr]
                feature_len = len(features_2)
                for k in range(feature_len):
                    sheet.cell(row=idx, column=k + 1).value = features_2[k]
                # features_2 = [name, lListId[i], lListId[j], lw, lh, lv, lw_2, lh_2, lv_2, ncc_2, sift_2, surf_2, orb_2, ls, ls_2, inter_distance, vr, sr]
                # for k in range(18):
                #     sheet.cell(row=idx, column=k + 1).value = features_2[k]
                idx += 1  
    wb.save(excelFile)

def generateOutput(excelFile, imgUrl):
    '''Generate output images (without generating dataset)
    @excelFile [in]: recording excel file
    @imgUrl [in]: folder contains input images
    '''
    left_dir = imgUrl + '\\left'
    left_dir_out = left_dir + '_out'
    right_dir = imgUrl + '\\right'
    right_dir_out = right_dir + '_out'
    if os.path.isdir(left_dir_out) is False:
        os.makedirs(left_dir_out)
    if os.path.isdir(right_dir_out) is False:
        os.makedirs(right_dir_out)
    imgList = next(os.walk(left_dir))[2]
    for name in imgList:
        lImg = left_dir + '\\' + name
        lOutDir = left_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        detectLEDRegion(lImg, lOutDir, excelFile)
        rImg = right_dir + '\\' + name
        rOutDir = right_dir_out + '\\' + os.path.splitext(name)[0] + '.txt'
        detectLEDRegion(rImg, rOutDir, excelFile)

def predictStereoMatching(features):
    '''Predict if 2 taillights are stereo matched or not (to be implemented later):
    @features [in]: input feature vector,
    @predict value [out]: two tailights are matched (1) or not (0)
    '''
    return 0

def predictTaillightPairing(features):
    '''Predict if 2 taillights are paired or not (to be implemented later):
    @features [in]: input feature vector,
    @predict value [out]: two tailights are paired (1) or not (0)
    '''
    return 0

def stereoMatching(leftImg, rightImg, excelFile):
    '''Stereo matching flow:
    @leftImg [in]: left image file name,
    @rightImg [in]: right image file name,
    @excelFile [in]: excel file,
    @matchList [out]: list of stereo matched taillights
    '''
    leftOup = os.path.splitext(ntpath.basename(leftImg))[0] + '_left.txt'
    rightOup = os.path.splitext(ntpath.basename(rightImg))[0] + '_right.txt'
    lOutDir = ntpath.dirname(leftImg) + leftOup
    rOutDir = ntpath.dirname(rightImg) + rightOup
    lList = detectLEDRegion(leftImg, leftDir, excelFile)
    rList = detectLEDRegion(rightImg, rOutDir, excelFile)
    lListId = lList.mGetLabelList()
    rListId = rList.mGetLabelList()
    isSkip = False
    matchList = list()#list of stereo matched taillights
    for lid in lListId:
        lbd, lw, lh, lv, ldn, ldnn, ls = lList.mGetRegionInfo(lid, utils.STEREO.LEFT)
        for rid in rListId:
            rbd, rw, rh, rv, rdn, rdnn, rs = rList.mGetRegionInfo(rid, utils.STEREO.RIGHT)
            ncc = lList.mGetNCCValue(lid, rList, rid)
            features = [lbd, lw, lh, lv, ldn, rbd, rw, rh, rv, rdn, ncc]
            predict = predictStereoMatching(features)
            if predict == 1:
                matchList.append((lid, rid))
                rListId.remove(rid)
                # isSkip = True
                # break
        # if(isSkip is True):
        #     isSkip = False
        #     continue
    return matchList

def taillightPairing(img, excelFile):
    '''Taillight pairing flow:
    @img [in]: image file name,
    @excelFile [in]: excel file,
    @pairList [out]: list of paired taillights
    '''
    oup = os.path.splitext(ntpath.basename(img))[0] + '.txt'
    outDir = ntpath.dirname(img) + oup
    lList =  detectLEDRegion(img, outDir, excelFile)
    lListId = lList.mGetLabelList()
    lLen = len(lListId)
    isSkip = False
    pairList = list()#list of paired taillights
    for i in range(lLen - 1):
        _, lw, lh, lv, _, _, ls = lList.mGetRegionInfo(lListId[i], utils.STEREO.LEFT)
        for j in range(i + 1, lLen):
            _, lw_2, lh_2, lv_2, _, _, ls_2 = lList.mGetRegionInfo(lListId[j], utils.STEREO.LEFT)
            inter_distance = lList.mGetInterDistance(lListId[i], lListId[j])
            ncc = lList.mGetNCCValue(lListId[i], lList, lListId[j], True)
            features = [lw, lh, lv, lw_2, lh_2, lv_2, ncc, inter_distance]
            predict = predictTaillightPairing(features)
            if predict == 1:
                pairList.append((lListId[i], lListId[j]))
                #isSkip = True
                break
        # if(isSkip is True):
        #     isSkip = False
        #     continue
    return pairList


# generateDataset('training_data_Img_2.xlsx', 'Dataset\\Img_2')
# generateDataset('training_data_Set_2.xlsx', 'Dataset\\Set_2')
generateMatchingDataset('Set_2\\matching_Set_2.xlsx', 'Dataset\\Set_2')
generatePairingDataset('Set_2\\pairing_left_Set_2.xlsx', 'Dataset\\Set_2\\left')
generatePairingDataset('Set_2\\pairing_right_Set_2.xlsx', 'Dataset\\Set_2\\right')
# im = cv2.imread('frame6963.jpg')
# thres = thresholdColor(im)
# labels = assignLabel(thres, im, 'result.txt')#, '')
# showResult('result.jpg', im, thres, labels)