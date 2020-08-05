import sys
import numpy as np
import cv2
import statistics 
import sklearn.preprocessing as pp
import utils
from skimage.feature import hog
from collections import OrderedDict
from enum import Enum
from openpyxl import load_workbook
from scipy import signal

#utils.feature_map = ['DBL': 1, 'SL': 2, 'VL': 3, 'DNL': 4, 'DBR': 1, 'SR': 2, 'VR': 3, 'DNR': 4]

#utils.feature_map = {'ImageID': 1, 'ID': 2, 'DB': 3, 'W': 4, 'H': 5, 'V': 6, 'DN': 7, 'DNN': 8, 'S': 9}#, 'HOG': 9, 'SIFT': 10, ''}#, 'DB': 1, 'SR': 2, 'VR': 3, 'DNR': 4]

idx = 2
isLeft = False

# utils.POS_INF = 999999
# utils.NEG_INF = -1
# utils.NUM_OF_BIN = 16

# utils.DESCRIPTOR_DSIFT = 0
# utils.DESCRIPTOR_HOG = 1

# utils.SIZE_THRES = 5
# utils.HEIGHT_THRES = float(3/4)
# utils.BRAKE_THRES = 5

# class utils.FD(Enum):
#     SIFT = 1
#     SURF = 2
#     ORB = 3
#     DENSE_SIFT = 4
#     DENSE_SURF = 5
#     DENSE_ORB = 6

# class utils.STEREO(Enum):
#     LEFT = 0
#     RIGHT = 1

# class utils.DIM(Enum):
#     HORIZONTAL = 0
#     VERTICAL = 1
#     BOTH = 2

class CRegion:
    '''To represent each LED region'''
    def __init__(self, _id):
        '''Init region with label/ID is _id:
        @id: label/id of the region'''
        self.pId = _id #region label/id
        self.pQuantity = 0 #number of points
        self.pPoints = OrderedDict() #key: point coordinate, value: point HSV value
        self.x_max = utils.NEG_INF #vertical max
        self.x_min = utils.POS_INF #vertical min
        self.y_max = utils.NEG_INF #horizontal max
        self.y_min = utils.POS_INF #horizontal min
        self.x = -1 #Oy-axis coordinate value
        self.y = -1 #Ox-axis coordinate value
        self.bf = cv2.BFMatcher(cv2.NORM_L1, crossCheck=False)

    def __str__(self):
        '''String representation of region'''
        st = ''
        for x in self.pPoints:
            st += str(x) + ':' + str(self.pPoints[x]) +'\n'
        return st

    def mGetId(self):
        '''Get region's label/ID'''
        return self.pId

    def mGetQuantity(self):
        '''Get the number of points in the region'''
        return self.pQuantity

    def mAddPoint(self, x_axis, y_axis, value):
        ''''Add new point to region:
        @x_axis: x-axis coord of the new point,
        @y_axis: y-axis coord of the new point,
        @value: HSV value of the new point'''
        self.pQuantity += 1
        self.pPoints[(x_axis, y_axis)] = value
        if x_axis < self.x_min:
            self.x_min = x_axis
        if x_axis > self.x_max:
            self.x_max = x_axis
        if y_axis < self.y_min:
            self.y_min = y_axis
        if y_axis > self.y_max:
            self.y_max = y_axis

    def mGetPoint(self, x_axis, y_axis):
        '''Get the HSV value of the point in region with specified coord:
        @x_axis: x-axis coord of the point,
        @y_axis: y-axis coord of the point'''
        if (x_axis, y_axis) in self.pPoints:
            return self.pPoints[(x_axis, y_axis)]
        return None

    def mGetPoints(self):
        '''Get the list of all points in the region'''
        return self.pPoints
    
    def mGetDenseKeypoints(self):
        '''Get the list of all region points as keypoint'''
        if self.mCheckBoundaryConstraint() is True:
            kp = [cv2.KeyPoint(i, j, 1) for (i, j) in self.pPoints]
            return kp
        return None

    def mGetDenseMBRKeypoints(self):
        '''Get the list of all points in MBR as keypoint'''
        kps = list()
        for i in range(self.x_min, self.x_max):
            for j in range(self.y_min, self.y_max):
                kps.append(cv2.KeyPoint(i, j, 1))
        return kps

    def mGetKeypoints(self, kp):#,prt=True):
        '''Extract keypoint list of region'''
        kp_list = list(filter(lambda x: x.pt[0] >= self.x_min and x.pt[0] <= self.x_max and x.pt[1] >= self.y_min and x.pt[1] <= self.y_max, kp))
        # if prt is False:
        #     f = open('debug.txt', 'w+')
        #     f.write(str(self) + '\n' + str(len(kp_list)) + '\n\n')
        #     f.close()
        return kp_list
    #def mGetROI(self, img):
    # Extract ROI from image

    def mGetCoords(self):
        '''Get the coordinate list of all points in the region'''
        pointCoords = []
        for point in self.pPoints:
            pointCoords.append(list(point))
        return pointCoords

    def mGetBoundaryCoord(self):
        '''Get coordinates of the MBR vertices of the region'''
        return [self.x_min, self.y_min, self.x_max, self.y_max]
        # pointCoords = np.array(self.mGetCoords())
        # x_max = np.amax(pointCoords[:, 0])
        # x_min = np.amin(pointCoords[:, 0])
        # y_max = np.amax(pointCoords[:, 1])
        # y_min = np.amin(pointCoords[:, 1])
        # return [x_min, y_min, x_max, y_max]

    def mGetBoundarySize(self):
        '''Get MBR size'''
        height = self.x_max - self.x_min
        width = self.y_max - self.y_min
        return [width, height]

    def mGetBoundaryArea(self):
        '''Get MBR area'''
        return (self.x_max - self.x_min) * (self.y_max - self.y_min)

    def mCheckBoundaryConstraint(self):
        '''Check the boundary constraint to validate object'''
        if self.x_min == utils.POS_INF or self.y_min == utils.POS_INF or self.x_max == utils.NEG_INF or self.y_max == utils.NEG_INF:#check initial state
            return False
        if self.x_min != self.x_max or self.y_min != self.y_max:
            return True
        return False

    def mGetCentroid(self):
        '''Get the region centroid (centroid of the minimum bounding rectangle)'''
        # size = self.mGetBoundarySize()
        # return [size[0] / 2, size[1] / 2]
        if self.x == -1:
            self.x = (self.x_max + self.x_min) / 2
        if self.y == -1:
            self.y = (self.y_max + self.y_min) / 2
        return [self.x, self.y]

    def mGetPos(self, dim=utils.DIM.BOTH):
        '''
        Get taillight's image coordinate (centroid of the MBR)
        dim [in]: dimension value want to get (default: both dimension)
        '''
        if self.x == -1:
            self.x = (self.x_max + self.x_min) / 2
        if self.y == -1:
            self.y = (self.y_max + self.y_min) / 2
        if dim == utils.DIM.HORIZONTAL:
            #print('Horizon')
            return self.y_min
            #return self.y
        elif dim == utils.DIM.VERTICAL:
            #print('VERTICAL ', self.y_min)
            return self.x_min
            #return self.x
        return [self.x_min, self.y_min]
        #return [self.x, self.y]
    
    # def mGetDistanceToBorder(self):
    # #Get the region's distance to left & right image border (corresponding to x_min, x_max)
    #     return self.x_min, self.x_max

    def mGetDistanceToBorder(self):#, width):#, stereo=utils.STEREO.LEFT):
        '''Get region's distance to image left border (for left image) or right border (for right image)
        @width [in]: image width
        @stereo [in]: point out if region in left/right image
        '''
        return self.y_min#horizontal min
        #return self.y
        # if stereo == utils.STEREO.LEFT:
        #     #print('Come in')
        #     return self.x_min
        # #width, _ = img.size()
        # #print('Come out')
        # return width - self.x_min
    
    #self-defined NCC function
    def mGetNCCValue_2(self, opponent, this_img, opp_img, isSameImg=False):
        '''
        Get NCC value with the opponent region
        @opponent [in]: opponent region
        @this_img [in]: image contains this region
        @opp_img [in]: image contains opponent region
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''
        [opp_x_min, opp_y_min, opp_x_max, opp_y_max] = opponent.mGetBoundaryCoord()
        this_img_norm = cv2.cvtColor(this_img, cv2.COLOR_BGR2GRAY)
        opp_img_norm = cv2.cvtColor(opp_img, cv2.COLOR_BGR2GRAY)
        this_img_norm = this_img_norm[self.x_min: self.x_max, self.y_min: self.y_max]
        opp_img_norm = opp_img_norm[opp_x_min: opp_x_max, opp_y_min: opp_y_max]
        if isSameImg is True:
            opp_img_norm = cv2.flip(opp_img_norm, 1)
        # this_img = np.asarray(this_img)
        # opp_img = np.asarray(opp_img)
        # this_norm = np.linalg.norm(this_img)
        # opp_norm = np.linalg.norm(opp_img)
        # this_row_sums = this_img.sum(axis=1)
        # this_img_new = this_img / this_row_sums[:, np.newaxis]
        # opp_row_sums = opp_img.sum(axis=1)
        # opp_img_new = opp_img / opp_row_sums[:, np.newaxis]
        #corr = signal.correlate2d(this_img, opp_img)
        #_, max_val, _, _ = cv2.minMaxLoc(corr)
        # max_val = np.max(corr)
        # return max_val / (this_norm * opp_norm)
        # this_img_m = this_img[this_pos_x: this_pos_x + width, this_pos_y : this_pos_y + height]
        # opp_img_m = this_img[opp_pos_x: opp_pos_x + width, opp_pos_y : opp_pos_y + height]
        # print(np.shape(this_img_m))
        # print(np.shape(opp_img_m))
        # res = cv2.matchTemplate(this_img_m, opp_img_m, cv2.TM_CCORR_NORMED)
        # _, max_val, _, _ = cv2.minMaxLoc(res)
        # return max_val
        this_mean = np.mean(this_img_norm)
        print(this_mean)
        opp_mean = np.mean(opp_img_norm)
        print(opp_mean)
        this_mean_sub = np.subtract(this_img_norm, this_mean)
        opp_mean_sub = np.subtract(opp_img_norm, opp_mean)
        this_norm = np.linalg.norm(this_mean_sub)
        opp_norm = np.linalg.norm(opp_mean_sub)
        #nominator = signal.correlate2d(this_img, opp_img)
        # this_std = np.std(this_img)
        # opp_std = np.std(opp_img)
        nominator = signal.correlate2d(this_mean_sub, opp_mean_sub)
        denominator = this_norm * opp_norm
        corr = np.divide(nominator, denominator)
        return np.max(corr)#nominator / denominator

    def mGetNCCValue(self, opponent, this_img, opp_img, isSameImg=False):
        '''
        Get NCC value with the opponent region
        @opponent [in]: opponent region
        @this_img [in]: image contains this region
        @opp_img [in]: image contains opponent region
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''
        bins = [16, 16, 16]#[50, 60]
        channels = [0, 1, 2]
        ranges = [0, 180, 0, 256, 0, 256]
        # hbin = 180
        # sbin = 256
        # vbin = 256
        [opp_x_min, opp_y_min, opp_x_max, opp_y_max] = opponent.mGetBoundaryCoord()
        # this_img = np.array(this_img)
        # opp_img = np.array(opp_img)
        this_img_norm = cv2.cvtColor(this_img, cv2.COLOR_BGR2HSV)
        opp_img_norm = cv2.cvtColor(opp_img, cv2.COLOR_BGR2HSV)     
        this_img_norm = this_img_norm[self.x_min: self.x_max, self.y_min: self.y_max, :]
        opp_img_norm = opp_img_norm[opp_x_min: opp_x_max, opp_y_min: opp_y_max, :]
        if isSameImg is True:
            opp_img_norm = cv2.flip(opp_img_norm, 1)
        this_hist = cv2.calcHist(this_img_norm, channels, None, bins, ranges, accumulate=False)
        opp_hist = cv2.calcHist(opp_img_norm, channels, None, bins, ranges, accumulate=False)
        return  cv2.compareHist(this_hist, opp_hist, 0)

    def mGetBhattaValue(self, opponent, this_img, opp_img, isSameImg=False):
        '''
        Get Bhatta value with the opponent region
        @opponent [in]: opponent region
        @this_img [in]: image contains this region
        @opp_img [in]: image contains opponent region
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''
        bins = [16, 16, 16]#[50, 60]
        channels = [0, 1, 2]
        ranges = [0, 180, 0, 256, 0, 256]
        # hbin = 180
        # sbin = 256
        # vbin = 256
        [opp_x_min, opp_y_min, opp_x_max, opp_y_max] = opponent.mGetBoundaryCoord()
        this_img_norm = cv2.cvtColor(this_img, cv2.COLOR_BGR2HSV)
        opp_img_norm = cv2.cvtColor(opp_img, cv2.COLOR_BGR2HSV)       
        this_img_norm = this_img_norm[self.x_min: self.x_max, self.y_min: self.y_max, :]
        opp_img_norm = opp_img_norm[opp_x_min: opp_x_max, opp_y_min: opp_y_max, :] 
        if isSameImg is True:
            opp_img_norm = cv2.flip(opp_img_norm, 1)
        this_hist = cv2.calcHist(this_img_norm, channels, None, bins, ranges, accumulate=False)
        opp_hist = cv2.calcHist(opp_img_norm, channels, None, bins, ranges, accumulate=False)
        return  cv2.compareHist(this_hist, opp_hist, 3)
        return np.max(corr)#nominator / denominator
    # def mGetFeatureDescriptor(self, img):
    # #'''Extract feature descriptor of region:
    # # @img: image of the input image
    # # '''
    #     region = img[self.x_min: self.x_max, self.y_min: self.y_max, :]        
    #     hsv = cv2.cvtColor(region, cv2.COLOR_BGR2HSV)
    #     try:
    #         h, s, v = cv2.split(hsv)
    #         kps = []
    #         des = np.array([])
    #         print('I am in\n')
    #         sift = cv2.xfeatures2d.SIFT_create()
    #         kp_h, des_h = sift.detectAndCompute(h, None)
    #         kps.append(kp_h)
    #         np.append(des, des_h)
    #         kp_s, des_s = sift.detectAndCompute(s, None)
    #         kps.append(kp_s)
    #         np.append(des, des_s)
    #         kp_v, des_v = sift.detectAndCompute(v, None)
    #         kps.append(kp_v)
    #         np.append(des, des_v)  
    #         return kps, des
    #     except:
    #         return None
        #region = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)         
        #print(region)
        # try:
        #     sift = cv2.xfeatures2d.SIFT_create()
        #     kp, des = sift.detectAndCompute(region, None)
        #     return kp, des
        # except:
        #     return None

    def mGetHOG(self, img, bin_n=utils.NUM_OF_BIN):
        '''
        Calculate HOG region descriptor
        @img [in]: input image
        @bin_n [n]: number of bins (default is 16)
        @hist [out]: region HOG (Histogram of Oriented Gradient)
        '''
        # x_min, y_min, x_max, y_max = self.mGetBoundaryCoord()
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        if self.mCheckBoundaryConstraint() is True:
            img = img[self.x_min: self.x_max, self.y_min: self.y_max, :]
            gx = cv2.Sobel(img, cv2.CV_32F, 1, 0)
            gy = cv2.Sobel(img, cv2.CV_32F, 0, 1)
            mag, ang = cv2.cartToPolar(gx, gy)
            bins = np.int32(bin_n * ang / (2 * np.pi)) #quantizing binvalues in (0...bin_n - 1)   
            #divide region into 4 sub-squares
            bin_cells = bins[:10, :10], bins[10:, :10], bins[:10, 10:], bins[10:, 10:]
            mag_cells = mag[:10, :10], mag[10:, :10], mag[:10, 10:], mag[10:, 10:]
            hists = [np.bincount(b.ravel(), m.ravel(), bin_n) for b, m in zip(bin_cells, mag_cells)]
            hist = np.hstack(hists) #hist is a 64-bit vector    
            return hist
        return None

    def mGetHOGInstant(self, img, bin_n= utils.NUM_OF_BIN):
        '''
        Calculate HOG region descriptor using built-in HOG function
        @img [in]: input image
        @bin_n [n]: number of bins (default is 16)
        @hist [out]: region HOG (Histogram of Oriented Gradient)
        '''
        # x_min, y_min, x_max, y_max = self.mGetBoundaryCoord()
        hist, _ = hog(img, orientations=bin_n, pixels_per_cell=(16, 16), cells_per_block=(1, 1), visualize=True, multichannel=True)   
        return hist 
    
    # def mCompareHOG(self, opponent, this_img, opp_img, isSameImg=False, bin_n= utils.NUM_OF_BIN):
    #     this_size = self.mGetBoundarySize()
    #     opp_size = opponent.mGetBoundarySize()
    #     width = max(this_size[0], opp_size[0])
    #     height = max(this_size[1], opp_size[1])
    #     if width == 0 or height == 0:
    #         #print('Access?')
    #         return 0
    #     this_pos_x = self.x_min
    #     this_pos_y = self.y_min
    #     [opp_pos_x, opp_pos_y] = opponent.mGetPos()
    #     this_img = cv2.cvtColor(this_img, cv2.COLOR_BGR2GRAY)
    #     opp_img = cv2.cvtColor(opp_img, cv2.COLOR_BGR2GRAY)
    #     #print(np.shape(this_img))
    #     this_img = this_img[this_pos_x: this_pos_x + this_size[1], this_pos_y: this_pos_y + this_size[0]]
    #     opp_img = opp_img[opp_pos_x: opp_pos_x + opp_size[1], opp_pos_y: opp_pos_y + opp_size[0]]
    #     this_hist = self.mGetHOGInstant(this_img)
    #     opp_hist = opponent.mGetHOGInstant(opp_img)

    def mGetDSIFT(self, img):#, kps=[]):
        '''
        Get dense SIFT descriptor of the region
        @img [in]: input image
        @kps [in]: list of keypoint to be used (default is dense keypoint)
        @dSIFT [out]: dense SIFT descriptor
        '''
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        sift = cv2.xfeatures2d.SIFT_create()
        kp = self.mGetDenseMBRKeypoints()
        # if not kps:
        #     #print('Hi')
        #     kp = self.mGetDenseMBRKeypoints()
        # else:
        #     #print('HELLO')
        #     kp = self.mGetKeypoints(kps)#, prt)
        _, dSIFT = sift.compute(img, kp)
        print('dSIFT shape: ', np.shape(dSIFT))
        dSIFT_norm = pp.normalize(dSIFT, norm='l2')
        return dSIFT_norm

    def mGetSIFT(self, img):
        '''
        Get SIFT descriptor of the region
        @img [in]: input image
        @dSIFT [out]: SIFT descriptor
        '''
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        #img = img[self.x_min: self.x_max, self.y_min: self.y_max]
        sift = cv2.xfeatures2d.SIFT_create()
        dSIFT = sift.detectAndCompute(img, None)
        print('SIFT shape: ', np.shape(dSIFT))
        dSIFT_norm = pp.normalize(dSIFT, norm='l2')
        return dSIFT_norm

    def mGetDORB(self, img):
        '''
        Get dense BRIEF descriptor of the region
        @img [in]: input image
        @des [out]: dense BRIEF descriptor
        '''
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        orb = cv2.ORB_create()
        kp = self.mGetDenseMBRKeypoints()
        #brief = cv2.xfeatures2d_BriefDescriptorExtractor()
        _, des = orb.compute(img, kp)
        return des    

    def mGetORB(self, img, kps=[]):
        '''
        Get ORB descriptor of the region
        @img [in]: input image
        @kps [in]: list of keypoint to be used (default is null - dense keypoint)
        @dSIFT [out]: ORB descriptor
        '''
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        orb = cv2.ORB_create()
        if not kps:
            kp = self.mGetDenseKeypoints()
        else:
            kp = self.mGetKeypoints(kps)
        _, dORB = orb.compute(img, kp)
        return dORB
    
    def mGetDSURF(self, img):#, kps=[]):
        '''
        Get dense SURF descriptor of the region
        @img [in]: input image
        @kps [in]: list of keypoint to be used (default is null - dense keypoint)
        @dSIFT [out]: dense SURF descriptor
        '''
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        surf = cv2.xfeatures2d.SURF_create()
        kp = self.mGetDenseMBRKeypoints()
        # if not kps:
        #     kp = self.mGetDenseMBRKeypoints()
        # else:
        #     kp = self.mGetKeypoints(kps)
        _, dSURF = surf.compute(img, kp)
        print('dSURF shape: ', np.shape(dSURF))
        dSURF_norm = pp.normalize(dSURF, norm='l2')
        return dSURF_norm

    def mGetSURF(self, img):#, kps=[]):
        '''
        Get SURF descriptor of the region
        @img [in]: input image
        @dSIFT [out]: SURF descriptor
        '''
        # img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # img = img[self.x_min: self.x_max, self.y_min: self.y_max]
        surf = cv2.xfeatures2d.SURF_create()
        dSURF = surf.detectAndCompute(img, None)
        dSURF_norm = pp.normalize(dSURF, norm='l2')
        return dSURF_norm

    # def mGetFeatureDescriptor(self, descriptor_type):
    #     if descriptor_type == utils.FD.SIFT or descriptor_type == utils.FD.DENSE_SIFT:
            
    def mGetRTable(self, img, descriptor):
        '''Get R-Table description
        img [in]: input image
        descriptor [in]: descriptor type (SIFT, SURF, ORB)
        des [out]: R-Table description'''
        des = []
        kp = self.pPoints
        rTable = OrderedDict()
        if descriptor == utils.FD.SIFT:
            des = self.mGetDSIFT(img)
        elif descriptor == utils.FD.SURF:
            des = self.mGetDSURF(img)
        else:
            des = self.mGetDORB(img)
        index = 0
        for (i, j) in kp:
            key = (abs(i - centroid[0]), j - centroid[1]) #create index
            rTable[key].append(dSIFT[index]) #append feature vector to corresponding entry on r-Table region descriptor
            index += 1
        return rTable
        # #centroid = self.mGetCentroid()
        # #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # #sift = cv2.xfeatures2d.SIFT_create()
        # #rTable = OrderedDict()
        # if self.mCheckBoundaryConstraint() is True:
        #     #img = img[self.x_min: self.x_max, self.y_min: self.y_max]#, :]
        #     #kp = [cv2.KeyPoint(x, y, 1) for y in range(self.x_min, self.x_max) 
        #                             #for x in range(self.y_min, self.y_max)]
        #     kp = [cv2.KeyPoint(i, j, 1) for (i, j) in self.pPoints] #initialize list of "keypoint" (all region points)
        #     #dSIFT = sift.compute(img, kp) #compute keypoint feature feature vectors
        #     # index = 0
        #     # for (i, j) in self.pPoints:
        #     #     key = (abs(i - centroid[0]), j - centroid[1]) #create index
        #     #     rTable[key].append(dSIFT[index]) #append feature vector to corresponding entry on r-Table region descriptor
        #     #     index += 1
        #     # return rTable
        #     return kp#, dSIFT
        # return None
    
    def mGetMatches(self, this_img, opponent_region, opponent_img, descriptor, kps):
        '''Determine matches between 2 regions
        this_img [in]: the image contain this region
        opponent_region [in]: the region to be matched
        opponent_img [in]: the image contain opponent_region
        descriptor [in]: descriptor type (SIFT, SURF, ORB)
        des [out]: R-Table description'''
        this_descriptor = []
        opponent_descriptor = []
        if descriptor == utils.FD.SIFT or descriptor == utils.FD.DENSE_SIFT:
            this_descriptor = self.mGetDSIFT(this_img, kps)
            opponent_descriptor = opponent_region.mGetDSIFT(opponent_img, kps)
            #print(opponent_region)
        elif descriptor == utils.FD.SURF or descriptor == utils.FD.DENSE_SURF:
            this_descriptor = self.mGetDSURF(this_img, kps)
            opponent_descriptor = opponent_region.mGetDSURF(opponent_img, kps) 
        elif descriptor == utils.FD.ORB:
            this_descriptor = self.mGetORB(this_img, kps)
            opponent_descriptor = opponent_region.mGetORB(opponent_img, kps)                      
        else:
            this_descriptor = self.mGetDORB(this_img)
            opponent_descriptor = opponent_region.mGetDORB(opponent_img)
        matches = self.bf.match(this_descriptor, opponent_descriptor)
        matches = sorted(matches, key=lambda x: x.distance)
        _matches = opponent_region.bf.match(opponent_descriptor, this_descriptor)
        _matches = sorted(_matches, key=lambda x: x.distance)
        #print(matches)
        return matches, _matches

    def mGetDistanceToRegion(self, region, dim=utils.DIM.HORIZONTAL):
        '''Calculate distance from this region to another region (distance of their top-left pos)
        region [in]: region to be calculated distance
        dist [out]: distance'''
        this_pos = self.mGetPos()
        reg_pos = region.mGetPos()
        if dim == utils.DIM.BOTH:
            dist = np.linalg.norm(np.subtract(this_pos, reg_pos))
        elif dim == utils.DIM.HORIZONTAL: #horizontal distance
            dist = abs(this_pos[1] - reg_pos[1])
        else:
            dist == abs(this_pos[0] - reg_pos[0])
        return dist

    def mGetFeatureDescriptor(self, img, descriptor_type=utils.FD.DENSE_SIFT):
        '''Get region's feature (dense keypoint) descriptor
        @img [in]: input image
        @descriptor_type [in]: feature descriptor type (SIFT/SURF/ORB)
        @feature_descriptor [out]: feature descriptor
        '''
        #print('Image shape: ', np.shape(img))
        if descriptor_type == utils.FD.DENSE_SIFT:
            feature_descriptor = self.mGetDSIFT(img)
        elif descriptor_type == utils.FD.DENSE_SURF:
            feature_descriptor = self.mGetDSURF(img)
        elif descriptor_type == utils.FD.DENSE_ORB:
            feature_descriptor = self.mGetDORB(img)
        elif descriptor_type == utils.FD.SIFT:
            feature_descriptor = self.mGetSIFT(img)
        elif descriptor_type == utils.FD.SURF:
            feature_descriptor = self.mGetSURF(img)
        elif descriptor_type == utils.FD.ORB:
            feature_descriptor = self.mGetORB(img)
        return feature_descriptor
        #return pp.normalize(feature_descriptor)#, norm='l2')

    def mCalcDescriptorDiff(self, opponent, this_img, opp_img, descriptor_type=utils.FD.DENSE_SIFT, isSameImg=False):
        '''
        Calculate difference between 2 regions' feature descriptors
        @opp [in]: opponent region
        @this_img [in]: input image of this region
        @opp_img [in]: input image of opponent region
        @descriptor_type [in]: feature descriptor type (SIFT/SURF/ORB)
        '''
        [opp_x_min, opp_y_min, opp_x_max, opp_y_max] = opponent.mGetBoundaryCoord()
        this_img = cv2.cvtColor(this_img, cv2.COLOR_BGR2GRAY)
        opp_img = cv2.cvtColor(opp_img, cv2.COLOR_BGR2GRAY)
        this_img_norm = this_img[self.x_min: self.x_max, self.y_min: self.y_max]
        opp_img_norm = opp_img[opp_x_min: opp_x_max, opp_y_min: opp_y_max]
        if isSameImg is True:
            opp_img_norm = cv2.flip(opp_img_norm, 1)
        this_descriptor = self.mGetFeatureDescriptor(this_img_norm, descriptor_type)
        opp_descriptor = opponent.mGetFeatureDescriptor(opp_img_norm, descriptor_type)
        print('this descriptor: ', this_descriptor)
        print('opp descriptor: ', opp_descriptor)
        matches = self.bf.match(this_descriptor, opp_descriptor)
        matches = sorted(matches, key=lambda x: x.distance)
        match_dist = np.zeros(10)
        idx = 0
        # if isSameImg is False:
        #     print('Different images!')
        # else:
        #     print('Same images!')
        for match in matches[:10]:
            match_dist[idx] = match.distance
            print(match.distance)
            idx += 1
        #match_dict_norm = pp.normalize(match_dist)
        return np.mean(match_dist)

    #def mCalcDistHorizontal(self, region)

class CList:
    # '''To manage all LED regions'''
    
    instance = None#static instance

    def __init__(self, _img, _origin):#_width, _height):
        '''
        Init region list with label/ID is _id:
        @_img: labelled image,
        @_origin: original image
        '''
        self.pRegions = OrderedDict()
        self.pQuantity = 0
        self.pImg = _img
        self.pOrigin = _origin
        _width, _height, _ = self.pImg.shape
        self.width = _width
        self.height = _height
        self.pLabelMat = np.full((self.width, self.height), np.inf)#avoid the case label is zero a.k.a, background  

    def __str__(self):
        '''String representation of region'''
        st = str(self.pQuantity) + '\n'
        quantity = 0
        for x in self.pRegions:
            st += str(quantity) + ':' + str(self.pRegions[x].mGetQuantity())+ '\n' + str(self.pRegions[x]) + '\n'
            quantity += 1
        return st

    def mGetList(self):
        '''Get all regions in list'''
        return self.pRegions

    def mGetQuantity(self):
        '''Get number of regions'''
        return self.pQuantity

    def mGetLabelMat(self):
        '''Get label matrix'''
        return self.pLabelMat

    def mGetLabelList(self):
        '''Get list of region label'''    
        return list(self.pRegions.keys())
        #fg.close()

    def mGetImg(self):
        '''
        Get image of taillight list
        '''
        return self.pImg

    def mGetOrigin(self):
        '''
        Get original image
        '''
        return self.pOrigin

    def mAppend(self, _id, x_axis, y_axis, value):
        '''Append new point to a specified region
        @_id: id of the region to add point,
        @x_axis: x-axis coord of the new point,
        @y_axis: y-axis coord of the new point,
        @value: HSV value of the new point'''
        if _id not in self.pRegions:
            self.pRegions[_id] = CRegion(_id)
            self.pQuantity += 1
        self.pRegions[_id].mAddPoint(x_axis, y_axis, value)
        self.pLabelMat[x_axis, y_axis] = _id

    def mGetBoundaries(self):
        '''Get list of each region's boundary'''
        boundaries = []
        label = []
        for x in self.pRegions:
            boundaries.append(self.pRegions[x].mGetBoundaryCoord())
            label.append(x)
        return np.array(boundaries), label

    def mMergeRegions(self, old_id, new_id):
        '''Merge one region to another region:
        @old_id: id of the merging region,
        @new_id: id of the merged region'''
        if new_id != old_id and old_id in self.pRegions:
            pts = self.pRegions[old_id].mGetPoints()
            for axes in pts:
                self.mAppend(new_id, axes[0], axes[1], pts[axes])
                self.pLabelMat[axes[0], axes[1]] = new_id
            self.pQuantity -= 1
            del self.pRegions[old_id]

    def mGetLabel(self, x_axis, y_axis):
        '''Get id/label of the specified point:
        @x_axis: x-axis coord of the point,
        @y_axis: y-axis coord of the point'''
        if x_axis < 0 or x_axis > self.width - 1 or y_axis < 0 or y_axis > self.height - 1:#out of image range
            return np.inf
        return self.pLabelMat[x_axis, y_axis]

    def mGetRegionDescriptor(self):
        '''Get list of every region descriptor:'''
        kps = OrderedDict()
        des = OrderedDict()
        f = open("Output\\Text\\Descriptor.txt", "w+")
        for _id in self.pRegions:
            if self.pRegions[_id].mCheckBoundary() is True:
                if self.pRegions[_id].mGetRegionKeypoints(self.pImg) is not None:
                    kps[_id], des[_id] = self.pRegions[_id].mGetRegionKeypoints(self.pImg)
                    f.write(str(self.pRegions[_id].mGetBoundaryCoord()) + ':' + str(np.shape(kps[_id])) + '\n')
                else:
                    print('Error occurs')
        return kps, des

    def mGetRegion(self, hid):
        '''Get taillight region from list
        @hid [in]: region id
        '''
        return self.pRegions[hid]

    def mGetNCCValue(self, hid, oppList, oid, isSameImg=False):
        '''
        Get NCC value between 2 regions
        @hid [in]: region's id
        @oppList [in]: opponent's region list
        @oid [in]: opponent region id
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''
        oppRegion = oppList.mGetRegion(oid)
        oppImg = oppList.mGetOrigin()
        return self.pRegions[hid].mGetNCCValue(oppRegion, self.pOrigin, oppImg, isSameImg), self.pRegions[hid].mGetNCCValue_2(oppRegion, self.pOrigin, oppImg, isSameImg)

    def mGetBhattaValue(self, hid, oppList, oid, isSameImg=False):
        '''
        Get Bhatta value between 2 regions
        @hid [in]: region's id
        @oppList [in]: opponent's region list
        @oid [in]: opponent region id
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''
        oppRegion = oppList.mGetRegion(oid)
        oppImg = oppList.mGetOrigin()
        return self.pRegions[hid].mGetBhattaValue(oppRegion, self.pOrigin, oppImg, isSameImg)

    def mGetDescriptionDiff(self, hid, oppList, oid, descriptor_type=utils.FD.DENSE_SIFT, isSameImg=False):
        '''
        Get description difference value between 2 regions
        @hid [in]: region's id
        @oppList [in]: opponent's region list
        @oid [in]: opponent region id
        @descriptor_type [in]: descriptor type (SIFT/SURF/ORV)
        @isSameImg [in]: if true, get NCC value of regions from the same image and otherwise
        '''    
        oppRegion = oppList.mGetRegion(oid)
        oppImg = oppList.mGetOrigin()    
        return self.pRegions[hid].mCalcDescriptorDiff(oppRegion, self.pOrigin, oppImg, descriptor_type, isSameImg)

    def mSuppress(self):
        '''Remove outliers (brake lamps, noise...)'''
        noise = [_id for _id in self.pRegions if self.pRegions[_id].mCheckBoundaryConstraint is False\
            #suppress tiny regions
            or (self.pRegions[_id].mGetBoundarySize()[0] <= utils.SIZE_THRES)\
            or (self.pRegions[_id].mGetBoundarySize()[1] <= utils.SIZE_THRES)\
            #suppress region relatively high from road
            or self.pRegions[_id].mGetBoundaryCoord()[0] <= (utils.HEIGHT_THRES * self.width)\
            #suppress potential brake lamp
            or self.pRegions[_id].mGetBoundarySize()[0] >= utils.BRAKE_THRES * self.pRegions[_id].mGetBoundarySize()[1]]
        for _id in noise:
            #print('Remove noise')
            del self.pRegions[_id]
            self.pQuantity -= 1
    
    def mGetListKeypoints(self, img):
        #desc = []
        kps = []
        #img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        for _id in self.pRegions:
            k = self.pRegions[_id].mGetDSift(img)
            #desc.extend(d)
            kps.extend(k)
            #des.append(self.pRegions[_id].mGetDSift())
        sift = cv2.xfeatures2d.SIFT_create()
        #print(np.shape(kps))
        _, desc = sift.compute(img, kps)
        #print(np.shape(desc))
        return desc, kps

    def mGetPair(self, descriptor=utils.FD.SIFT):
        '''List all potential pairs of taillights'''
        f = open('Output\\Text\\match.txt', 'w+')
        kp = []
        self.pImg = cv2.cvtColor(self.pImg, cv2.COLOR_RGB2GRAY)
        if descriptor == utils.FD.SIFT:
            sift = cv2.xfeatures2d.SIFT_create()
            kp = sift.detect(self.pImg, None)
        elif descriptor == utils.FD.ORB:
            orb = cv2.ORB_create()
            kp = orb.detect(self.pImg, None)
        elif descriptor == utils.FD.SURF:
            surf = cv2.xfeatures2d.SURF_create()
            kp = surf.detect(self.pImg, None)
        for key, value in self.pRegions.items():
            for _key, _value in self.pRegions.items():
                if key != _key:
                    matches = value.mGetMatches(self.pImg, _value, self.pImg, descriptor, kp)
                    f.write('(' + str(key) + ',' + str(_key) + '):' + str(len(matches)) + '\n')
                    f.write('(' + str(_key) + ',' + str(key) + '):' + str(len(_matches)) + '\n')
        f.close()

    def mGetStereoHorizontalDistance(self, hid, stereo=utils.STEREO.LEFT):
        '''
        Get the horizontal distance from the specified region to its nearest right (for left img) or left (for right img) neighbor
        @hid [in]: region id
        @stereo [in]: specify which image the object in (left/right img)
        '''
        # dist_left = np.inf
        # dist_right = np.inf
        dist = np.inf
        for _id in self.pRegions:
            if _id != hid and self.pRegions[_id].mGetPos(utils.DIM.VERTICAL) > self.pRegions[hid].mGetPos(utils.DIM.VERTICAL)\
                and ((stereo == utils.STEREO.LEFT and self.pRegions[_id].mGetPos(utils.DIM.HORIZONTAL) > self.pRegions[hid].mGetPos(utils.DIM.HORIZONTAL))\
                or (stereo == utils.STEREO.RIGHT and self.pRegions[_id].mGetPos(utils.DIM.HORIZONTAL) < self.pRegions[hid].mGetPos(utils.DIM.HORIZONTAL))):
                distCand = self.pRegions[hid].mGetDistanceToRegion(self.pRegions[_id], utils.DIM.HORIZONTAL)
                if distCand < dist:
                    dist = distCand
        #     if _if != hid:
        #         distCand = self.pRegions[hid].mGetDistanceToRegion(self.pRegions[_id], utils.DIM.HORIZONTAL)
        #         if self.pRegions[_id].mGetPos(utils.DIM.HORIZONTAL) > self.pRegions[hid].mGetPos(utils.DIM.HORIZONTAL) and distCand < dist_right:
        #             dist_right = distCand
        #         if self.pRegions[_id].mGetPos(utils.DIM.HORIZONTAL) < self.pRegions[hid].mGetPos(utils.DIM.HORIZONTAL) and distCand < dist_left:
        #             dist_left = distCand                
        # return dist_left, dist_right
        return dist

    def mGetHorizontalDistanceToNearestRegion(self, hid):
        '''
        Get the horizontal distance from the specified region to its nearest neighbor
        @hid [in]: region id
        '''
        dist = np.inf
        for _id in self.pRegions:
            if _id != hid:
                distCand = self.pRegions[hid].mGetDistanceToRegion(self.pRegions[_id], utils.DIM.HORIZONTAL)
                if distCand < dist:
                    dist = distCand
        return dist

    def mGetInterDistance(self, hid, oid, dim=utils.DIM.HORIZONTAL):
        '''
        Get inter-distance between 2 regions in list
        @hid [in]: first region
        @oid [in]: second region
        @dim [in]: dimension (horizontal/vertical/both)
        '''
        return self.pRegions[hid].mGetDistanceToRegion(self.pRegions[oid], dim) 

    #def mGetNCCValue(self, )
    def mWriteToFile(self, desname, excelFile=None):
        '''Get number of regions
        @desname [in]: output file
        @excelFile [in]: excel file to record data
        '''
        global idx
        global isLeft
        #self.mSuppress()
        
        f = open(desname, "w+")
        if excelFile is not None:
            wb = load_workbook(excelFile)
        if 'left' in desname:
            # if isLeft is True:
            #     isLeft = False
            #     idx = 2
            _stereo = utils.STEREO.LEFT
            if excelFile is not None:
                sheet = wb.worksheets[0]
        else:
            if isLeft is False:
                isLeft = True
                idx = 2
            _stereo = utils.STEREO.RIGHT
            if excelFile is not None:
                sheet = wb.worksheets[1]
        if excelFile is not None:
            sheet.cell(row=idx, column=utils.feature_map['ImageID']).value = desname
            sheet.cell(row=idx + 1, column=utils.feature_map['ImageID']).value = str(self.pQuantity)
        #print('File loaded')
        #index = 2
        for x in self.pRegions:
            # print('width: ', self.width)
            # print('height: ', self.height)
            #print('Write boundary')
            #fg.write(str(index) + ':' + str(self.pRegions[x].mGetQuantity()) + '\n')
            #f.write(str(self.pRegions[x].mGetBoundaryCoord()) + ':' + str(self.pRegions[x].mGetBoundarySize()) + '\n')
            f.write(str(x) + ':' + str(self.pRegions[x].mGetBoundaryCoord()) + ':' + str(self.pRegions[x].mGetBoundarySize()) + '\n')
            #print('Can access text')
            if excelFile is not None:
                sheet.cell(row=idx, column=utils.feature_map['ID']).value = str(x)
            #bd = self.pRegions[x].mGetDistanceToBorder(self.width, _stereo) / float(self.width)
                bd = self.pRegions[x].mGetDistanceToBorder() / float(self.height)
                sheet.cell(row=idx, column=utils.feature_map['DB']).value = str(bd)#str(self.pRegions[x].mGetDistanceToBorder(self.pImg, _stereo))
                [width, height] = self.pRegions[x].mGetBoundarySize()
                sheet.cell(row=idx, column=utils.feature_map['W']).value = str(width/ float(self.height))
                sheet.cell(row=idx, column=utils.feature_map['H']).value = str(height/ float(self.width))
                v = self.pRegions[x].mGetPos(utils.DIM.VERTICAL) / float(self.width)
            #print(x, desname, str(self.pRegions[x].mGetPos(utils.DIM.VERTICAL)))
                sheet.cell(row=idx, column=utils.feature_map['V']).value = str(v)#self.pRegions[x].mGetPos(utils.DIM.VERTICAL)
            #[dnl, dnr] = self.mGetStereoHorizontalDistance(x, _stereo) / float(self.height)
                dn = self.mGetStereoHorizontalDistance(x, _stereo) / float(self.height)
                if dn == np.inf:
                    dn = 1
            #dnl, dnr = self.mGetStereoHorizontalDistance(x) / float(self.height)
            # if dnl == np.inf:
            #     dnl = 1
            # if dnr == np.inf:
            #     dnr = 1
            # sheet.cell(row=idx, column=utils.feature_map['DNL']).value = str(dnl)#self.mGetStereoHorizontalDistance(x, _stereo)
            # sheet.cell(row=idx, column=utils.feature_map['DNR']).value = str(dnr)
                sheet.cell(row=idx, column=utils.feature_map['DN']).value = str(dn)
                dnn = self.mGetHorizontalDistanceToNearestRegion(x) / float(self.height)
                if dnn == np.inf:
                    dnn = 1
                sheet.cell(row=idx, column=utils.feature_map['DNN']).value = str(dnn)
                size = (width / float(self.height)) * (height/ float(self.width))
                sheet.cell(row=idx, column=utils.feature_map['S']).value = str(size)
                idx += 1
        f.close()
        if excelFile is not None:
            wb.save(excelFile)
    
    def mGetVerticalCentroidRatio(self, hid, oid):
        [this_x, _] = self.pRegion[hid].mGetCentroid()
        [opp_x, _] = self.pRegion[oid].mGetCentroid()
        [_, this_h] = self.pRegion[hid].mGetBoundarySize()
        [_, opp_h] = self.pRegion[oid].mGetBoundarySize()
        diff_x = abs(this_x - opp_x)
        diff_h = (this_h + opp_h) / 2
        if(diff_x <= diff_h):
            return 1
        return 0  

    def mGetRegionInfo(self, hid, stereo_info):
        '''Get information of the specified taillight region
        @hid [in]: taillight ID
        @stereo_info [in]: stereo information (left/right image)
        @bd [out]: distance to border
        @w [out]: width
        @h [out]: height
        @v [out]: vertical position
        @dn [out]: distance to nearest left (in right img) and right (in left img) neighbor
        @dnn [out]: distance to nearest neighbor
        @s [out]: region size
        '''
        bd = self.pRegions[hid].mGetDistanceToBorder() / float(self.height)
        [w, h] = self.pRegions[hid].mGetBoundarySize()
        w = w /  float(self.height)
        h = h / float(self.width)
        s = w * h
        v = self.pRegions[hid].mGetPos(utils.DIM.VERTICAL) / float(self.width)
        # dnl, dnr = self.mGetStereoHorizontalDistance(hid, stereo_info) / float(self.height)
        # if dnl == np.inf:
        #     dnl = 1
        # if dnr == np.inf:
        #     dnr = 1
        dn = self.mGetStereoHorizontalDistance(hid, stereo_info) / float(self.height)
        if dn == np.inf:
            dn = 1
        dnn = self.mGetHorizontalDistanceToNearestRegion(hid) / float(self.height)
        if dnn == np.inf:
            dnn = 1
        #centroid = self.mGetRegion(hid).mGetCentroid()
        return bd, w, h, v, dn, dnn, s
        #return str(bd), str(w), str(h), str(v), str(dn), str(dnn), str(s)#, centroid #str(dnl), str(dnr)

    #def mCalcFeatureDescriptor(self, )

    @staticmethod
    def mGetInstance(width, height):
        '''Get static instance width specified size of input image
        @width: image width
        @height: image height'''
        if not CList.instance:
            CList.instance = CList(width, height)
        return CList.instance
